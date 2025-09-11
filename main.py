from __future__ import annotations
from version import __version__ as APP_VERSION

import csv
import os
import re
import sys
import threading
import tkinter as tk
import webbrowser
from difflib import SequenceMatcher
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, ttk

import cv2
import fitz  # PyMuPDF
import numpy as np
import openpyxl  # для .xlsx
import pytesseract
import requests
from dotenv import load_dotenv, set_key
from PIL import Image, ImageTk

from auto_updater import AutoUpdater
from custom_logger import logger
from version import __version__  # Импортируем номер версии

EASYOCR_AVAILABLE = False
PADDLEOCR_AVAILABLE = False


# Глобальные переменные (лучше использовать класс для состояния)
DEFAULT_COORDINATES2 = {
    "X_START": 20,
    "Y_START": 298,
    "X_END": 92,
    "Y_END": 345,
    "REGEX_PATTERN": r"^[A-Z]{3}U\d{7}$",
}

global table_frame, debug_mode

original_page_image = None
pdf_path = None
image_display = None
rect_id = None
text_output = None
# pdf = None
pdf_doc = None
scale_percent = 100  # Масштаб для обработки координат
ENV_FILE = ".env"

# reader = None
# Добавьте в раздел глобальных переменных
ocr_engine = "Tesseract"  # По умолчанию
ocr_reader = None  # Для хранения инициализированного ридера EasyOCR/PaddleOCR
current_page = 0
canvas_scale = 1.0

table_entries = []  # глобальная переменная — таблица как список словарей
expected_containers = []  # список контейнеров из XLS
selected_areas = []
recognition_results = []  # Будет хранить словари с результатами для каждой страницы
last_click_time = 0
DOUBLE_CLICK_DELAY = 300  # Задержка для двойного клика в миллисекундах

# статус-бар: переменные
status_page_var = None
status_zoom_var = None
status_size_var = None
status_msg_var  = None

current_theme = "light"  # по умолчанию


# 1-based для Treeview identify_column / bbox:
NUMBER_COL = 1
EXPECTED_COL = 2
INVOICE_COL = 3
RECOGNIZED_COL = 4
MATCH_COL = 5
SCORE_COL = 6

# Удобные 0-based индексы для массива values:
IDX_NUMBER = NUMBER_COL - 1
IDX_EXPECTED = EXPECTED_COL - 1
IDX_INVOICE = INVOICE_COL - 1
IDX_RECOGNIZED = RECOGNIZED_COL - 1
IDX_MATCH = MATCH_COL - 1
IDX_SCORE = SCORE_COL - 1


# Настройка логирования
# logging.basicConfig(
#     filename="app.log",
#     level=logging.DEBUG,
#     format="%(asctime)s - %(levelname)s - %(message)s",
# )

def status_set(*, page=None, total=None, zoom=None, size=None, msg=None):
    """Обновляет элементы статус-бара; передавай только то, что меняется."""
    if page is not None and total is not None and status_page_var is not None:
        status_page_var.set(f"Стр: {page}/{total}")
    if zoom is not None and status_zoom_var is not None:
        status_zoom_var.set(f"Масштаб: {zoom}")
    if size is not None and status_size_var is not None:
        status_size_var.set(f"Размер: {size}")
    if msg is not None and status_msg_var is not None:
        status_msg_var.set(msg)


def apply_minimal_theme2(root, theme="light"):
    """
    Лёгкий, лаконичный стиль для Tkinter/ttk с авто-темой и аккуратным Treeview.

    :param root: Tkinter root window
    :param theme: "light" или "dark"
    """
    import tkinter as tk
    from tkinter import ttk

    try:
        import sv_ttk
        sv_ttk.set_theme(theme)
        print("[THEME] sv_ttk применена:", theme)  # ← добавь
        use_sv = True
    except Exception as e:
        print("[THEME] sv_ttk не работает:", e)  # ← и это
        use_sv = False

    # 1) Пробуем современную тему sv_ttk (Sun Valley)
    try:
        import sv_ttk  # pip install sv-ttk
        sv_ttk.set_theme(theme)  # 'light' или 'dark'
        use_sv = True
    except Exception:
        use_sv = False
        style = ttk.Style(root)
        style.theme_use("clam")

        if theme == "dark":
            # TODO: добавить полноценную ручную темную тему при необходимости
            BG = "#2E2E2E"
            FG = "#F5F5F5"
            ACCENT = "#3B82F6"
            MUTED = "#9CA3AF"
            SEL_BG = "#374151"
        else:
            # Светлая тема по умолчанию
            BG = "#F7F7F9"
            FG = "#111827"
            ACCENT = "#2563EB"
            MUTED = "#6B7280"
            SEL_BG = "#DBEAFE"

        style.configure(".", background=BG, foreground=FG, font=("Segoe UI", 10))
        style.configure("TFrame", background=BG)
        style.configure("TLabel", background=BG, foreground=FG)
        style.configure("TButton", padding=8, relief="flat")
        style.map("TButton",
                  background=[("active", "#E5E7EB")],
                  relief=[("pressed", "sunken")])

        style.configure("Toolbar.TFrame", background=BG)
        style.configure("ToolSep.TFrame", background="#E5E7EB")

        style.configure("TEntry", padding=6)
        style.configure("TCombobox", padding=6)

        style.configure("Treeview",
                        borderwidth=0,
                        rowheight=28,
                        font=("Segoe UI", 10))
        style.configure("Treeview.Heading",
                        font=("Segoe UI Semibold", 10),
                        foreground=MUTED)
        style.map("Treeview",
                  background=[("selected", SEL_BG)],
                  foreground=[("selected", FG)])

    # 2) Универсальные «хелперы» для зебры и тулбара
    def style_treeview_stripes(tree):
        try:
            tree.tag_configure("oddrow", background="#F3F4F6")
            for i, iid in enumerate(tree.get_children("")):
                if i % 2 == 1:
                    tree.item(iid, tags=("oddrow",))
        except Exception:
            pass

    def build_toolbar(parent, *widgets):
        bar = ttk.Frame(parent, style="Toolbar.TFrame")
        bar.pack(side="top", fill="x")
        for w in widgets:
            w.pack(in_=bar, side="left", padx=6, pady=8)
        sep = ttk.Frame(parent, style="ToolSep.TFrame", height=1)
        sep.pack(side="top", fill="x")
        return bar

    # 3) Сохраним хелперы на root (для внешнего доступа)
    root._style_helpers = {
        "style_treeview_stripes": style_treeview_stripes,
        "build_toolbar": build_toolbar,
        "sv_ttk": use_sv
    }

def apply_minimal_theme(root, theme="light"):
    """Минималистичный светлый/тёмный стиль для ttk и sv_ttk (если доступен)."""
    global current_theme
    import tkinter as tk
    from tkinter import ttk

    try:
        import sv_ttk
        sv_ttk.set_theme(theme)
        print("[DEBUG] sv_ttk.set_theme успешно:", theme)  # ← ЭТО ВАЖНО
        use_sv = True
    except Exception as e:
        use_sv = False
        style = ttk.Style(root)
        style.theme_use("clam")
        print("[DEBUG] sv_ttk ошибка:", e)

        # Цветовая палитра
        if theme == "dark":
            BG = "#1F2937"  # фоновый
            FG = "#E5E7EB"  # текст
            ACCENT = "#3B82F6"
            MUTED = "#9CA3AF"
            SEL_BG = "#374151"
        else:
            BG = "#F7F7F9"
            FG = "#111827"
            ACCENT = "#2563EB"
            MUTED = "#6B7280"
            SEL_BG = "#DBEAFE"

        style.configure(".", background=BG, foreground=FG, font=("Segoe UI", 10))
        style.configure("TFrame", background=BG)
        style.configure("TLabel", background=BG, foreground=FG)
        style.configure("TButton", padding=8, relief="flat")
        style.map("TButton",
                  background=[("active", "#E5E7EB")],
                  relief=[("pressed", "sunken")])

        style.configure("Toolbar.TFrame", background=BG)
        style.configure("ToolSep.TFrame", background="#E5E7EB")

        style.configure("TEntry", padding=6)
        style.configure("TCombobox", padding=6)

        style.configure("Treeview",
                        borderwidth=0,
                        rowheight=28,
                        font=("Segoe UI", 10))
        style.configure("Treeview.Heading",
                        font=("Segoe UI Semibold", 10),
                        foreground=MUTED)
        style.map("Treeview",
                  background=[("selected", SEL_BG)],
                  foreground=[("selected", FG)])

    def style_treeview_stripes(tree):
        try:
            tree.tag_configure("oddrow", background="#F3F4F6")
            for i, iid in enumerate(tree.get_children("")):
                if i % 2 == 1:
                    tree.item(iid, tags=("oddrow",))
        except Exception:
            pass

    def build_toolbar(parent, *widgets):
        bar = ttk.Frame(parent, style="Toolbar.TFrame")
        bar.pack(side="top", fill="x")
        for w in widgets:
            w.pack(in_=bar, side="left", padx=6, pady=8)
        sep = ttk.Frame(parent, style="ToolSep.TFrame", height=1)
        sep.pack(side="top", fill="x")
        return bar

    root._style_helpers = {
        "style_treeview_stripes": style_treeview_stripes,
        "build_toolbar": build_toolbar,
        "sv_ttk": use_sv
    }


current_theme = "light"  # по умолчанию

def toggle_theme():
    global current_theme
    current_theme = "dark" if current_theme == "light" else "light"
    apply_minimal_theme(root, current_theme)


def toggle_theme2():
    """Переключатель темы: если есть sv_ttk — используем его; иначе ручной фоллбэк."""
    from tkinter import ttk
    helpers = getattr(root, "_style_helpers", {})
    if helpers.get("sv_ttk"):
        try:
            import sv_ttk
            sv_ttk.toggle_theme()
            return
        except Exception:
            pass

    # --- Фоллбэк для обычного ttk (clam) ---
    style = ttk.Style(root)
    current = getattr(root, "_theme_mode", "light")

    if current == "light":
        # тёмная
        BG = "#111827"; FG = "#E5E7EB"; MUTED = "#9CA3AF"; SEL_BG = "#374151"
        TB_BG = "#0F172A"; SEP = "#1F2937"
        style.configure(".", background=BG, foreground=FG, font=("Segoe UI", 10))
        style.configure("TFrame", background=BG)
        style.configure("TLabel", background=BG, foreground=FG)
        style.configure("TButton", padding=8)
        style.configure("Toolbar.TFrame", background=TB_BG)
        style.configure("ToolSep.TFrame", background=SEP)
        style.configure("Treeview", background=BG, fieldbackground=BG, foreground=FG, rowheight=28)
        style.configure("Treeview.Heading", foreground=MUTED)
        style.map("Treeview", background=[("selected", SEL_BG)], foreground=[("selected", FG)])
        root._theme_mode = "dark"
    else:
        # светлая
        BG = "#F7F7F9"; FG = "#111827"; MUTED = "#6B7280"; SEL_BG = "#DBEAFE"
        TB_BG = BG; SEP = "#E5E7EB"
        style.configure(".", background=BG, foreground=FG, font=("Segoe UI", 10))
        style.configure("TFrame", background=BG)
        style.configure("TLabel", background=BG, foreground=FG)
        style.configure("TButton", padding=8)
        style.configure("Toolbar.TFrame", background=TB_BG)
        style.configure("ToolSep.TFrame", background=SEP)
        style.configure("Treeview", background=BG, fieldbackground=BG, foreground=FG, rowheight=28)
        style.configure("Treeview.Heading", foreground=MUTED)
        style.map("Treeview", background=[("selected", SEL_BG)], foreground=[("selected", FG)])
        root._theme_mode = "light"


def is_similar_ratio(a, b):
    return SequenceMatcher(None, a, b).ratio()


def match_with_expected():
    global table_entries, expected_containers, tree

    # Получаем ожидаемые контейнеры из таблицы
    expected_containers = []
    for item in tree.get_children():
        values = tree.item(item, "values")
        if len(values) > 1 and values[1]:  # values[1] - столбец "Контейнер из XLS"
            expected_containers.append(values[1])

    if not expected_containers:
        messagebox.showwarning("Ошибка", "Нет данных для сопоставления в столбце 'Контейнер из XLS'")
        return

    for entry in table_entries:
        recognized = entry.get("recognized", "")
        if not recognized:
            continue

        best_match = ""
        best_score = 0.0

        for expected in expected_containers:
            if not expected:
                continue
            score = is_similar_ratio(recognized, expected)
            if score > best_score:
                best_score = score
                best_match = expected

        # Обновляем строку в таблице
        values = list(tree.item(entry["item_id"], "values"))
        values[4] = best_match  # Совпадение
        values[5] = f"{best_score:.2f}"  # Коэффициент
        tree.item(entry["item_id"], values=values)

        # Обновляем цвет строки
        if best_score == 1.0:  # Полное совпадение
            tree.tag_configure("exact_match", background="#a8e6a8")  # Светло-зеленый
            tree.item(entry["item_id"], tags=("exact_match",))
        elif best_score > 0:  # Любое другое совпадение
            tree.tag_configure("partial_match", background="#fff8a8")  # Светло-желтый
            tree.item(entry["item_id"], tags=("partial_match",))
        else:  # Нет совпадения
            tree.tag_configure("no_match", background="#ffaaaa")  # Светло-красный
            tree.item(entry["item_id"], tags=("no_match",))


def safe_execute(func):
    """Декоратор для оборачивания функций с обработкой ошибок и логированием."""

    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            logger.error(f"Ошибка в функции {func.__name__}: {e}", exc_info=True)
            messagebox.showerror("Ошибка", f"Произошла ошибка: {e}")

    return wrapper


def set_tesseract_path():
    global text_output
    # Список возможных путей установки Tesseract
    possible_paths = [
        Path("C:/Program Files/Tesseract-OCR/tesseract.exe"),
        Path("C:/Program Files (x86)/Tesseract-OCR/tesseract.exe"),
        Path.home() / "Tesseract-OCR/tesseract.exe",
        Path.home() / "AppData/Local/Tesseract-OCR/tesseract.exe",
        Path.home() / "AppData/Local/Programs/Tesseract-OCR/tesseract.exe",
    ]
    # Проверка пути в папке Programs

    # Проверка всех возможных путей
    for path in possible_paths:
        if path.exists():
            pytesseract.pytesseract.tesseract_cmd = str(path)
            logger.debug(f"Путь для Tesseract установлен: {path}")
            return

    # Если ни один путь не найден
    logger.info("Tesseract не найден. Пожалуйста, установите его по ссылке:")
    logger.info("Ссылка на проект: https://github.com/UB-Mannheim/tesseract/wiki")
    logger.info("Дистрибутив https://github.com/UB-Mannheim/tesseract/releases")
    insert_link(text_output, "Ссылка на проект:", "https://github.com/UB-Mannheim/tesseract/wiki")
    insert_link(text_output, "Дистрибутив:", "https://github.com/UB-Mannheim/tesseract/releases")


# Функция для создания выходной папки
def create_output_directory(input_file_path):
    try:
        base_name = Path(input_file_path).stem
        output_dir = Path(f"{base_name}_out")
        output_dir.mkdir(parents=True, exist_ok=True)
        return str(output_dir)  # Возвращаем строку для совместимости
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось создать выходную папку: {e}")
        return None

def draw_selection():
    global rect_id, x_start, y_start, x_end, y_end, canvas2, cropped_image, canvas2_scale

    if rect_id:
        canvas.delete(rect_id)
    rect_id = canvas.create_rectangle(x_start, y_start, x_end, y_end, outline="red", width=2)

    update_coordinates_entry()

    if page_image:
        # Вырезаем область с оригинальными координатами (без учета текущего масштаба)
        inverse_scale = 1 / scale_factor
        x1 = int(x_start * inverse_scale)
        y1 = int(y_start * inverse_scale)
        x2 = int(x_end * inverse_scale)
        y2 = int(y_end * inverse_scale)

        cropped_image = original_page_image.crop((x1, y1, x2, y2))

        # Масштабируем до 200% по умолчанию
        canvas2_scale = 1.0
        width = int(cropped_image.width * canvas2_scale)
        height = int(cropped_image.height * canvas2_scale)
        scaled_img = cropped_image.resize((width, height), Image.LANCZOS)

        # Отображаем на canvas2
        canvas2.delete("all")
        canvas2.image = ImageTk.PhotoImage(scaled_img)
        canvas2.create_image(0, 0, anchor=tk.NW, image=canvas2.image)
        canvas2.config(scrollregion=canvas2.bbox(tk.ALL))


def unload_pdf():
    global pdf_doc, current_page, page_image, image_display
    global original_page_image, selected_areas, total_pages
    global scale_factor, last_scale_factor

    try:
        if pdf_doc:
            pdf_doc.close()
            pdf_doc = None
    except Exception as e:
        logger.warning(f"Не удалось закрыть PDF: {e}")

    # Сброс параметров
    current_page = 0
    page_image = None
    original_page_image = None
    image_display = None
    selected_areas = []
    total_pages = 0
    scale_factor = 1.0
    last_scale_factor = 1.0

    # Очистка холста
    if canvas:
        canvas.delete("all")
        canvas.config(scrollregion=(0, 0, 0, 0))


def select_pdf():
    global pdf_path, current_page, entry_pdf_path
    try:
        file_path = filedialog.askopenfilename(filetypes=[("PDF файлы", "*.pdf")])
        if file_path:
            unload_pdf()  # ← сбрасываем предыдущий
            pdf_path = file_path
            entry_pdf_path.delete(0, tk.END)
            entry_pdf_path.insert(0, file_path)
            current_page = 0
            load_page()

    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось выбрать или загрузить PDF: {e}")
        logger.warning("Ошибка выбора PDF")


# Функция загрузки и отображения страницы PDF
@safe_execute
def load_page():
    global image_display, label_page_number, label_page_size, label_scale
    global pdf_doc, current_page, page_image, scale_factor
    global page_width2, page_height2, canvas, total_pages
    global original_page_image, last_scale_factor, pdf_path, tree

    if not pdf_path:
        return

    try:
        if pdf_doc is None:
            # recognition_results = []  # Очищаем предыдущие результаты
            pdf_doc = fitz.open(pdf_path)
            # Создаем строки таблицы по количеству страниц
            build_table_from_pdf(pdf_doc)

        current_page = max(0, min(current_page, pdf_doc.page_count - 1))
        page = pdf_doc.load_page(current_page)
        total_pages = pdf_doc.page_count

        pix = page.get_pixmap(dpi=200)
        original_page_image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

        canvas_height = canvas.winfo_height()
        image_height = original_page_image.height
        scale_factor = (scale_percent / 100) * (canvas_height / image_height)
        last_scale_factor = scale_factor

        scaled_width = int(original_page_image.width * scale_factor)
        scaled_height = int(original_page_image.height * scale_factor)
        page_image = original_page_image.resize((scaled_width, scaled_height), Image.LANCZOS)

        image_display = ImageTk.PhotoImage(image=page_image)
        if canvas:
            canvas.delete("all")
            canvas.create_image(0, 0, anchor=tk.NW, image=image_display)
            canvas.config(scrollregion=canvas.bbox(tk.ALL))

        title = (
            f"Распознавание текста из PDF - Страница {current_page + 1}/{total_pages} - "
            f"Координаты: ({x_start}, {y_start}) -> ({x_end}, {y_end})"
        )
        root.title(title)
        draw_selection()

    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось загрузить страницу: {e}")
        logger.warning("Ошибка при загрузке страницы")


# Функция для выбора координат области
def define_coordinates(event):
    global x_start, y_start, rect_id, scale_percent, scale_factor
    # Координаты события с учетом масштаба
    x_start = int(event.x)
    y_start = int(event.y)
    logger.info(scale_factor)
    logger.info(scale_percent)
    logger.info(event.x, event.y)
    logger.info(x_start, y_start)

    # Удаляем старое выделение, если оно есть
    if rect_id:
        canvas.delete(rect_id)

    # Создаем новый прямоугольник, начальная точка
    rect_id = canvas.create_rectangle(x_start, y_start, x_start, y_start, outline="red", width=2)
    update_coordinates_entry()


def draw_rectangle(event):
    global rect_id, x_start, y_start
    x_end, y_end = event.x, event.y
    if rect_id is not None:
        canvas.coords(rect_id, x_start, y_start, x_end, y_end)


def finish_coordinates(event):
    global x_start, y_start, x_end, y_end, cropped_image_display, cropped_image
    global canvas2, label_coordinates, total_pages

    x_end, y_end = event.x, event.y

    # Проверяем, чтобы координаты были корректны
    if x_end < x_start:
        x_start, x_end = x_end, x_start
    if y_end < y_start:
        y_start, y_end = y_end, y_start

    root.title(
        f"Распознавание текста из PDF - Страница {current_page + 1}/{total_pages} - Координаты:"
        f" ({x_start}, {y_start}) -> ({x_end}, {y_end})"
    )

    canvas.coords(rect_id, x_start, y_start, x_end, y_end)
    selected_areas.clear()
    selected_areas.append((rect_id, x_start, y_start, x_end, y_end))

    update_coordinates_entry()

    if page_image:
        try:
            cropped_image = page_image.crop((x_start, y_start, x_end, y_end))
            cropped_image_display = ImageTk.PhotoImage(image=cropped_image)
            canvas2.create_image(0, 0, anchor=tk.NW, image=cropped_image_display)
            canvas2.image = cropped_image_display
        except Exception as e:
            messagebox.showerror("Ошибка", f"Неверные координаты выделения: {e}")


# Функция обновления метки координат
def update_coordinates_label():
    global x_start, y_start, x_end, y_end, label_coordinates
    if x_start is not None and y_start is not None and x_end is not None and y_end is not None:
        label_coordinates.config(
            text=f"Координаты: x={x_start}, y={y_start}, width={x_end - x_start}, height={y_end - y_start}"
        )


def update_coordinates_entry():
    global x_start, y_start, x_end, y_end, coordinates_entry
    if x_start is not None and y_start is not None and x_end is not None and y_end is not None:
        coordinates_text = f"{x_start},{y_start},{x_end},{y_end}"
        coordinates_entry.delete(0, tk.END)
        coordinates_entry.insert(0, coordinates_text)


# Функция для проверки и форматирования распознанного текста
def format_extracted_text(text, i):
    # Удаление всех символов, кроме английских букв и цифр
    cleaned_text = re.sub(r"[^A-Za-z0-9]", "", text).upper()

    # Проверка формата: 4 буквы (в верхнем регистре) + 7 цифр
    if re.match(r"^[A-Z]{4}\d{7}$", cleaned_text):
        return cleaned_text

    # Замена неправильных символов на '@'
    formatted_text = ""
    for i, char in enumerate(cleaned_text):
        if i < 4 and not char.isalpha():
            formatted_text += "@"
        elif i < 4 and char.islower():
            formatted_text += char.upper()
        elif i >= 4 and not char.isdigit():
            formatted_text += "@"
        else:
            formatted_text += char

    # Проверка длины результата и обрезка лишнего
    formatted_text = formatted_text[:11]  # Убедиться, что длина не превышает 11 символов

    # Дополнение '@', если строка короче
    while len(formatted_text) < 11:
        formatted_text += "@"

    return formatted_text


def build_table_from_pdf(pdf_doc):
    global table_entries, tree

    # Очищаем таблицу
    for item in tree.get_children():
        tree.delete(item)

    table_entries = []

    # Создаем строки по количеству страниц
    for i in range(pdf_doc.page_count):
        item_id = tree.insert(
            "",
            tk.END,
            values=(
                i + 1,  # № страницы
                "",     # Контейнер из XLS (пока пусто)
                "",     # invoice (накладная из XLS) ← НОВОЕ
                "",     # Распознанный контейнер
                "",     # Совпадение
                "",     # Коэффициент
            ),
        )
        # ДОБАВИЛИ xls_id
        table_entries.append({
            "index": i + 1, "item_id": item_id,
            "code": "", "recognized": "", "xls_id": ""
        })


def update_table_from_entries(table_frame_ref):
    global table_entries

    # Очистка предыдущего содержимого таблицы
    for widget in table_frame_ref.grid_slaves():
        if int(widget.grid_info()["row"]) > 0:
            widget.destroy()

    for i, entry in enumerate(table_entries, start=1):
        # Ячейка "№"
        tk.Label(table_frame_ref, text=str(entry["index"]), borderwidth=1, relief=tk.RIDGE).grid(
            row=i, column=0, sticky="nsew"
        )

        # Ячейка "Контейнер из XLS" — может быть пустой
        tk.Label(table_frame_ref, text=entry.get("code", ""), borderwidth=1, relief=tk.RIDGE).grid(
            row=i, column=1, sticky="nsew"
        )

        # Ячейка "Контейнер распознанный"
        lbl_recognized = tk.Label(table_frame_ref, text=entry.get("recognized", ""), borderwidth=1, relief=tk.RIDGE)
        lbl_recognized.grid(row=i, column=2, sticky="nsew")
        entry["label_recognized"] = lbl_recognized

        # Ячейка "Совпадение"
        lbl_match = tk.Label(table_frame_ref, text=entry.get("match", ""), borderwidth=1, relief=tk.RIDGE)
        lbl_match.grid(row=i, column=3, sticky="nsew")
        entry["label_match"] = lbl_match

        # Ячейка "Коэффициент"
        lbl_score = tk.Label(table_frame_ref, text=entry.get("score", ""), borderwidth=1, relief=tk.RIDGE)
        lbl_score.grid(row=i, column=4, sticky="nsew")
        entry["label_score"] = lbl_score


# Функция для выполнения длительной задачи в потоке
def start_recognition_thread():
    threading.Thread(target=start_recognition, daemon=True).start()


def start_recognition():
    global selected_areas, pdf_doc, last_scale_factor

    if not pdf_doc:
        messagebox.showwarning("Нет документа", "Пожалуйста, выберите PDF-файл.")
        return

    # Если нет выделенной области, но есть координаты из .env
    if not selected_areas and all(v is not None for v in [x_start, y_start, x_end, y_end]):
        selected_areas = [(None, x_start, y_start, x_end, y_end)]
        draw_selection()  # Визуализируем область

    if not selected_areas:
        messagebox.showwarning("Нет выделения", "Пожалуйста, выделите область на холсте.")
        return

    try:
        area = selected_areas[0]
        _, x1, y1, x2, y2 = area
        coords = (x1, y1, x2, y2)
        engine = ocr_engine_var.get().lower()

        for page_num in range(pdf_doc.page_count):
            recognized_text = recognize_area(pdf_doc, page_num, coords, engine)
            formatted_text = format_extracted_text(recognized_text, page_num + 1)

            # Обновляем таблицу
            if page_num < len(table_entries):
                table_entries[page_num]["recognized"] = formatted_text
                item_id = table_entries[page_num]["item_id"]
                current_values = list(tree.item(item_id, "values"))
                current_values[3] = formatted_text  # recognized column
                tree.item(item_id, values=current_values)

        messagebox.showinfo("Готово", f"Распознано {pdf_doc.page_count} страниц.")

    except Exception as e:
        logger.warning("Ошибка при распознавании всех страниц")
        messagebox.showerror("Ошибка", f"Произошла ошибка при распознавании: {e}")


def save_results(btn):
    """Запускает сохранение результатов в отдельном потоке"""
    btn.config(state=tk.DISABLED)
    threading.Thread(target=_save_results_worker, args=(btn,), daemon=True).start()


def _save_results_worker2(btn):
    """Функция-рабочий для сохранения результатов (UI через root.after)."""
    global pdf_doc, table_entries, debug_mode, recognition_results

    # Все диалоги/окна — только через root.after
    if not pdf_doc:
        root.after(0, lambda: messagebox.showerror("Ошибка", "PDF документ не загружен"))
        root.after(0, lambda: btn.config(state=tk.NORMAL))
        return

    if not table_entries:
        root.after(0, lambda: messagebox.showerror("Ошибка", "Нет данных для сохранения"))
        root.after(0, lambda: btn.config(state=tk.NORMAL))
        return

    ui = {"win": None, "bar": None}

    def _open_progress():
        progress = tk.Toplevel(root)
        progress.title("Сохранение...")
        w, h = 300, 100
        root.update_idletasks()
        x = root.winfo_x() + (root.winfo_width() - w) // 2
        y = root.winfo_y() + (root.winfo_height() - h) // 2
        progress.geometry(f"{w}x{h}+{x}+{y}")
        progress.transient(root)
        progress.resizable(False, False)
        tk.Label(progress, text="Идет сохранение результатов").pack(pady=10)
        bar = ttk.Progressbar(progress, mode="indeterminate")
        bar.pack(fill="x", padx=20, pady=5)
        bar.start()
        ui["win"] = progress
        ui["bar"] = bar

    def _close_progress_ok(output_dir):
        if ui.get("bar"):
            try: ui["bar"].stop()
            except Exception: pass
        if ui.get("win"):
            try: ui["win"].destroy()
            except Exception: pass
        messagebox.showinfo("Сохранено", f"Результаты сохранены в папку:\n{output_dir}")
        btn.config(state=tk.NORMAL)

    def _close_progress_err(msg):
        if ui.get("bar"):
            try: ui["bar"].stop()
            except Exception: pass
        if ui.get("win"):
            try: ui["win"].destroy()
            except Exception: pass
        messagebox.showerror("Ошибка", msg)
        btn.config(state=tk.NORMAL)

    # открыть окно прогресса в главном потоке
    root.after(0, _open_progress)

    try:
        # куда сохраняем: рядом с исходным PDF (оставляю твою схему)
        output_dir = Path(entry_pdf_path.get()).parent

        for i, entry in enumerate(table_entries):
            page_num = entry["index"] - 1
            page = pdf_doc.load_page(page_num)
            pix = page.get_pixmap(dpi=200)
            page_image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

            recognized = (entry.get("recognized") or "").strip()
            match = (tree.item(entry["item_id"], "values")[4] or "").strip()  # "Совпадение"
            filename = match if match else recognized if recognized else f"page_{i + 1}"

            # префикс из колонки 3 Excel:
            prefix = (entry.get("xls_id") or "").strip()
            if prefix:
                filename = f"{prefix}_{filename}"

            output_file = Path(output_dir) / filename
            page_image.save(f"{output_file}.jpg")

            if debug_mode.get():
                # сохранить кроп, если действительно есть
                if 'cropped_image' in globals() and cropped_image is not None:
                    try:
                        cropped_image.save(f"{output_file}_cropped.jpg")
                    except Exception as e_crop:
                        logger.error(f"Не удалось сохранить вырезку: {e_crop}", exc_info=True)

                # сохранить _info.txt, если есть результат распознавания
                if i < len(recognition_results):
                    try:
                        result = recognition_results[i]
                        info_file = Path(f"{output_file}_info.txt")
                        with info_file.open("w", encoding="utf-8") as f:
                            f.write(f"Страница: {result.get('page')}\n")
                            f.write(f"Координаты: {result.get('coords')}\n")
                            f.write(f"Движок OCR: {result.get('engine')}\n")
                            f.write("\n--- Исходный текст ---\n")
                            f.write(result.get("raw_text", ""))
                            f.write("\n\n--- Форматированный текст ---\n")
                            f.write(result.get("formatted_text", ""))
                    except Exception as e_info:
                        logger.error(f"Не удалось сохранить _info.txt: {e_info}", exc_info=True)

        # закрыть прогресс и показать успех — в главном потоке
        root.after(0, lambda d=output_dir: _close_progress_ok(d))

    except Exception as e:
        logger.error(f"Ошибка сохранения: {e}", exc_info=True)
        # закрыть прогресс и показать ошибку — в главном потоке
        root.after(0, lambda: _close_progress_err(f"Ошибка при сохранении: {e!s}"))

def _save_results_worker(btn):
    """Функция-рабочий для сохранения результатов (UI через root.after).

    Имена файлов:
      - базовое имя: match → recognized → page_{i+1}
      - префикс накладной: ищем строку, где expected == (match|recognized),
        берём её invoice (приоритетно entry['xls_id'], иначе values[2])
        и формируем "{invoice}_{basename}".
    """
    global pdf_doc, table_entries, debug_mode, recognition_results, tree

    # Все диалоги/окна — только через root.after
    if not pdf_doc:
        root.after(0, lambda: messagebox.showerror("Ошибка", "PDF документ не загружен"))
        root.after(0, lambda: btn.config(state=tk.NORMAL))
        return

    if not table_entries:
        root.after(0, lambda: messagebox.showerror("Ошибка", "Нет данных для сохранения"))
        root.after(0, lambda: btn.config(state=tk.NORMAL))
        return

    ui = {"win": None, "bar": None}

    def _open_progress():
        progress = tk.Toplevel(root)
        progress.title("Сохранение...")
        w, h = 300, 100
        root.update_idletasks()
        x = root.winfo_x() + (root.winfo_width() - w) // 2
        y = root.winfo_y() + (root.winfo_height() - h) // 2
        progress.geometry(f"{w}x{h}+{x}+{y}")
        progress.transient(root)
        progress.resizable(False, False)
        tk.Label(progress, text="Идет сохранение результатов").pack(pady=10)
        bar = ttk.Progressbar(progress, mode="indeterminate")
        bar.pack(fill="x", padx=20, pady=5)
        bar.start()
        ui["win"] = progress
        ui["bar"] = bar

    def _close_progress_ok(output_dir):
        if ui.get("bar"):
            try:
                ui["bar"].stop()
            except Exception:
                pass
        if ui.get("win"):
            try:
                ui["win"].destroy()
            except Exception:
                pass
        messagebox.showinfo("Сохранено", f"Результаты сохранены в папку:\n{output_dir}")
        btn.config(state=tk.NORMAL)

    def _close_progress_err(msg):
        if ui.get("bar"):
            try:
                ui["bar"].stop()
            except Exception:
                pass
        if ui.get("win"):
            try:
                ui["win"].destroy()
            except Exception:
                pass
        messagebox.showerror("Ошибка", msg)
        btn.config(state=tk.NORMAL)

    # открыть окно прогресса в главном потоке
    root.after(0, _open_progress)

    try:
        # куда сохраняем: рядом с исходным PDF
        output_dir = Path(entry_pdf_path.get()).parent

        for i, entry in enumerate(table_entries):
            page_num = entry["index"] - 1
            page = pdf_doc.load_page(page_num)
            pix = page.get_pixmap(dpi=200)
            page_image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

            # --- 1) базовое имя: match -> recognized -> page_{i+1}
            row_values = list(tree.item(entry["item_id"], "values"))

            # match (с учётом новой колонки invoice)
            match_cell = (row_values[4] if len(row_values) > 4 and row_values[4] else "").strip()

            # recognized: сперва из entry (если уже записан), затем из таблицы (index 3)
            recognized_cell_in_table = (row_values[3] if len(row_values) > 3 and row_values[3] else "").strip()
            recognized = (entry.get("recognized") or recognized_cell_in_table or "").strip()

            basename = match_cell or recognized or f"page_{i + 1}"

            # --- 2) ищем накладную по expected == (match|recognized)
            invoice_prefix = ""
            search_key = (match_cell or recognized).strip()
            if search_key:
                for e in table_entries:
                    vals = list(tree.item(e["item_id"], "values"))
                    expected_val = (vals[1] if len(vals) > 1 and vals[1] else "").strip()
                    if expected_val == search_key:
                        # приоритетно из структуры записи, иначе из видимой таблицы (колонка invoice)
                        invoice_prefix = (e.get("xls_id") or (vals[2] if len(vals) > 2 else "") or "").strip()
                        if invoice_prefix:
                            break

            # --- 3) итоговое имя
            filename = f"{invoice_prefix}_{basename}" if invoice_prefix else basename

            # сохраняем
            output_file = Path(output_dir) / filename
            page_image.save(f"{output_file}.jpg")

            if debug_mode.get():
                # сохранить кроп, если действительно есть (глобальный для текущей сессии)
                if 'cropped_image' in globals() and cropped_image is not None:
                    try:
                        cropped_image.save(f"{output_file}_cropped.jpg")
                    except Exception as e_crop:
                        logger.error(f"Не удалось сохранить вырезку: {e_crop}", exc_info=True)

                # сохранить _info.txt, если есть результат распознавания
                if i < len(recognition_results):
                    try:
                        result = recognition_results[i]
                        info_file = Path(f"{output_file}_info.txt")
                        with info_file.open("w", encoding="utf-8") as f:
                            f.write(f"Страница: {result.get('page')}\n")
                            f.write(f"Координаты: {result.get('coords')}\n")
                            f.write(f"Движок OCR: {result.get('engine')}\n")
                            f.write("\n--- Исходный текст ---\n")
                            f.write(result.get("raw_text", ""))
                            f.write("\n\n--- Форматированный текст ---\n")
                            f.write(result.get("formatted_text", ""))
                    except Exception as e_info:
                        logger.error(f"Не удалось сохранить _info.txt: {e_info}", exc_info=True)

        # закрыть прогресс и показать успех — в главном потоке
        root.after(0, lambda d=output_dir: _close_progress_ok(d))

    except Exception as e:
        logger.error(f"Ошибка сохранения: {e}", exc_info=True)
        # закрыть прогресс и показать ошибку — в главном потоке
        root.after(0, lambda: _close_progress_err(f"Ошибка при сохранении: {e!s}"))



def run_ocr_in_thread(image, **kwargs):
    """Запуск OCR в отдельном потоке с передачей параметров."""
    threading.Thread(target=enhanced_recognition, args=(image,), kwargs=kwargs).start()


@safe_execute
def enhanced_recognition(
    image,
    use_grayscale=True,
    use_median_blur=True,
    use_thresholding=False,
    use_clahe=True,
    use_resize=True,
    use_deskew=False,
    use_noise_removal=True,
    use_morphological_ops=False,
    use_channel_extraction=False,
    channel="blue",
    use_edge_preprocessing=True,  # новый параметр
):
    """Расширенная функция распознавания текста с поддержкой EasyOCR и настройками включения этапов обработки."""

    # Преобразование в оттенки серого
    if use_grayscale:
        image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Предобработка через Canny после размытия
    if use_edge_preprocessing:
        blur = cv2.GaussianBlur(image, (5, 5), 0)
        image = cv2.Canny(blur, 50, 150)

    # Применение медианной фильтрации для удаления шума
    if use_median_blur:
        image = cv2.medianBlur(image, 3)

    # Применение бинаризации и пороговой обработки
    if use_thresholding:
        _, image = cv2.threshold(image, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    # Применение CLAHE для повышения контраста
    if use_clahe:
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        image = clahe.apply(image)

    # Масштабирование изображения для лучшего распознавания
    if use_resize:
        image = cv2.resize(image, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)

    # Коррекция наклона (deskew)
    if use_deskew:
        coords = np.column_stack(np.where(image > 0))
        angle = cv2.minAreaRect(coords)[-1]
        angle = -(90 + angle) if angle < -45 else -angle
        (h, w) = image.shape[:2]
        center = (w // 2, h // 2)
        M = cv2.getRotationMatrix2D(center, angle, 1.0)
        image = cv2.warpAffine(image, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)

    # Удаление шума с помощью морфологических операций
    if use_noise_removal:
        image = cv2.medianBlur(image, 3)

    # Морфологические операции для удаления мелких артефактов
    if use_morphological_ops:
        kernel = np.ones((1, 1), np.uint8)
        image = cv2.dilate(image, kernel, iterations=1)
        image = cv2.erode(image, kernel, iterations=1)

    # Извлечение цветового канала (если включено)
    if use_channel_extraction:
        if channel == "blue":
            image = image[:, :, 0]
        elif channel == "green":
            image = image[:, :, 1]
        elif channel == "red":
            image = image[:, :, 2]

    # extracted_text = pytesseract.image_to_string(image, lang="eng").strip()
    extracted_text = pytesseract.image_to_string(image, lang="eng").strip().upper()

    return extracted_text


def convert_coords_to_pdf(coords_canvas, scale_factor):
    x1, y1, x2, y2 = coords_canvas
    return (
        int(x1 / scale_factor),
        int(y1 / scale_factor),
        int(x2 / scale_factor),
        int(y2 / scale_factor),
    )


def extract_text_by_coords(page_num, coords_canvas):
    if not pdf_doc:
        logger.info("PDF-документ не загружен.")
        return ""

    coords_pdf = convert_coords_to_pdf(coords_canvas, last_scale_factor)

    try:
        page = pdf_doc.load_page(page_num)
        pix = page.get_pixmap(dpi=300)
        pil_img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        cropped = pil_img.crop(coords_pdf)
        return recognize_with_selected_engine(cropped, engine)
    except Exception as e:
        logger.info(f"Ошибка при OCR страницы {page_num + 1}: {e}")
        return ""


def check_image():
    global current_page, selected_areas, recognition_results

    if not pdf_doc:
        messagebox.showwarning("Нет файла", "Пожалуйста, выберите PDF-файл.")
        return

    # Если нет выделенной области, но есть координаты из .env
    if not selected_areas and all(v is not None for v in [x_start, y_start, x_end, y_end]):
        selected_areas = [(None, x_start, y_start, x_end, y_end)]
        draw_selection()  # Визуализируем область

    if not selected_areas:
        messagebox.showwarning("Нет выделения", "Пожалуйста, выделите область на холсте.")
        return

    try:
        area = selected_areas[0]
        _, x1, y1, x2, y2 = area
        coords = (x1, y1, x2, y2)
        engine = ocr_engine_var.get().lower()

        recognized_text = recognize_area(pdf_doc, current_page, coords, engine)
        formatted_text = format_extracted_text(recognized_text, current_page + 1)

        # Обновляем таблицу
        if current_page < len(table_entries):
            table_entries[current_page]["recognized"] = formatted_text
            tree.item(
                table_entries[current_page]["item_id"],
                values=(current_page + 1, table_entries[current_page]["code"], invoice, formatted_text, "", ""),
            )

        # Выводим в текстовое поле
        text_output.delete(1.0, tk.END)
        if current_page < len(recognition_results):
            result = recognition_results[current_page]
            # text_output.insert(tk.END, f"=== Страница {current_page + 1} ===\n")
            # text_output.insert(tk.END, f"Координаты: {result['coords']}\n")
            # text_output.insert(tk.END, f"Движок: {result['engine']}\n")
            # text_output.insert(tk.END, "\n--- Исходный текст ---\n")
            # text_output.insert(tk.END, result['raw_text'])
            # text_output.insert(tk.END, "\n\n--- Форматированный текст ---\n")
            # text_output.insert(tk.END, result['formatted_text'])
            logger.info(f"=== Страница {current_page + 1} ===\n")
            logger.info(f"Координаты: {result['coords']}\n")
            logger.info(f"Движок: {result['engine']}\n")
            logger.info("\n--- Исходный текст ---\n")
            logger.info(result["raw_text"])
            logger.info("\n\n--- Форматированный текст ---\n")
            logger.info(result["formatted_text"])

    except Exception as e:
        logger.error(f"Ошибка при распознавании страницы {current_page + 1}: {e}", exc_info=True)
        messagebox.showerror("Ошибка", f"Ошибка при распознавании: {e}")


def extract_area_image_from_pdf(pdf_doc, page_index, coords, dpi=200):
    """Вырезает область изображения из PDF по координатам"""
    x_start, y_start, x_end, y_end = coords

    # Проверяем корректность координат
    if x_end <= x_start or y_end <= y_start:
        raise ValueError("Некорректные координаты области выделения")

    page = pdf_doc.load_page(page_index)
    pix = page.get_pixmap(dpi=dpi)
    page_image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

    # Пересчёт координат из canvas → оригинальное изображение
    inverse_scale = 1 / last_scale_factor
    x0 = int(x_start * inverse_scale)
    y0 = int(y_start * inverse_scale)
    x1 = int(x_end * inverse_scale)
    y1 = int(y_end * inverse_scale)

    # Проверяем, чтобы координаты не выходили за границы изображения
    x0 = max(0, min(x0, page_image.width - 1))
    y0 = max(0, min(y0, page_image.height - 1))
    x1 = max(1, min(x1, page_image.width))
    y1 = max(1, min(y1, page_image.height))

    if x1 <= x0 or y1 <= y0:
        raise ValueError("Некорректные координаты после масштабирования")

    cropped = page_image.crop((x0, y0, x1, y1))
    if cropped.size[0] == 0 or cropped.size[1] == 0:
        raise ValueError("Выделенная область имеет нулевой размер")

    return cv2.cvtColor(np.array(cropped), cv2.COLOR_RGB2BGR)


def recognize_with_selected_engine(image, engine):
    """Распознает текст с использованием выбранного движка OCR"""
    engine = engine.lower().strip()

    if engine == "tesseract":
        return pytesseract.image_to_string(image, lang="eng").strip()

    if engine == "easyocr" and EASYOCR_AVAILABLE:
        results = ocr_reader.readtext(image, detail=0, paragraph=False)
        return " ".join(results).strip()

    if engine == "paddleocr" and PADDLEOCR_AVAILABLE:
        results = ocr_reader.predict(img=image, cls=True)
        if results and isinstance(results[0], list):
            return " ".join([line[1][0] for line in results[0]]).strip()
        return ""

    # Fallback на Tesseract, если движок неизвестен или недоступен
    return pytesseract.image_to_string(image, lang="eng").strip()


def recognize_area(pdf_doc, page_index, coords, engine):
    global recognition_results

    try:
        cropped_image = extract_area_image_from_pdf(pdf_doc, page_index, coords)
        if cropped_image is None or cropped_image.size == 0:
            return ""

        recognized_text = recognize_with_selected_engine(cropped_image, engine)
        formatted_text = format_extracted_text(recognized_text, page_index + 1)

        # Сохраняем результаты
        result = {
            "page": page_index + 1,
            "raw_text": recognized_text,
            "formatted_text": formatted_text,
            "coords": coords,
            "engine": engine,
        }

        # Обновляем или добавляем запись
        if len(recognition_results) > page_index:
            recognition_results[page_index] = result
        else:
            recognition_results.append(result)

        # Вывод в консоль (можно закомментировать)
        logger.info(f"Страница {page_index + 1}: {formatted_text}")

        return formatted_text

    except Exception as e:
        logger.error(f"Ошибка при распознавании области: {e}", exc_info=True)
        return ""


# Функция для перехода к следующей странице
def next_page():
    global current_page
    current_page += 1
    load_page()


# Функция для перехода к предыдущей странице
def prev_page():
    global current_page
    current_page -= 1
    load_page()

# Функция для обновления поля с дефолтными координатами
def set_default_coordinates(coordinates_entry):
    coordinates_text = (f"{DEFAULT_COORDINATES2['X_START']},{DEFAULT_COORDINATES2['Y_START']},"
                        f"{DEFAULT_COORDINATES2['X_END']},{DEFAULT_COORDINATES2['Y_END']}")
    # coordinates_entry.delete(0, tk.END)
    coordinates_entry.insert(0, coordinates_text)

# Функция для проверки формата координат
def validate_coordinates_format(coordinates_text):
    parts = coordinates_text.split(",")
    return len(parts) == 4 and all(part.isdigit() for part in parts)

def zoom_canvas(event):
    global canvas_scale, canvas, original_page_image

    if not original_page_image:
        return

    # Позиция мыши в canvas
    mouse_x = canvas.canvasx(event.x)
    mouse_y = canvas.canvasy(event.y)

    # Зум фактор
    zoom_factor = 1.1 if event.delta > 0 else 0.9
    new_scale = canvas_scale * zoom_factor

    # Ограничим масштаб от 10% до 500%
    new_scale = max(0.1, min(new_scale, 5.0))
    if new_scale == canvas_scale:
        return

    canvas_scale = new_scale

    # Размер нового изображения
    new_width = int(original_page_image.width * canvas_scale)
    new_height = int(original_page_image.height * canvas_scale)
    scaled_image = original_page_image.resize((new_width, new_height), Image.LANCZOS)

    # Обновляем изображение
    canvas.delete("all")
    img_tk = ImageTk.PhotoImage(scaled_image)
    canvas.image = img_tk  # Сохраняем ссылку
    canvas.create_image(0, 0, anchor=tk.NW, image=img_tk)
    canvas.config(scrollregion=(0, 0, new_width, new_height))

    # Смещение скроллбаров так, чтобы под курсором оставалась та же точка
    canvas.xview_moveto((mouse_x * zoom_factor - event.x) / new_width)
    canvas.yview_moveto((mouse_y * zoom_factor - event.y) / new_height)



def zoom_canvas2(event):
    global canvas2_scale, cropped_image

    if not cropped_image:
        return

    # Получаем координаты мыши относительно canvas2
    mouse_x = canvas2.canvasx(event.x)
    mouse_y = canvas2.canvasy(event.y)

    # Устанавливаем коэффициент масштабирования
    zoom_factor = 1.1 if event.delta > 0 else 0.9
    new_scale = canvas2_scale * zoom_factor

    # Ограничения масштаба
    new_scale = max(0.5, min(new_scale, 5.0))
    if new_scale == canvas2_scale:
        return

    canvas2_scale = new_scale

    # Новые размеры изображения
    new_width = int(cropped_image.width * canvas2_scale)
    new_height = int(cropped_image.height * canvas2_scale)

    # Масштабируем изображение
    scaled_img = cropped_image.resize((new_width, new_height), Image.LANCZOS)

    # Обновляем canvas2
    canvas2.delete("all")
    canvas2.image = ImageTk.PhotoImage(scaled_img)
    canvas2.create_image(0, 0, anchor=tk.NW, image=canvas2.image)
    canvas2.config(scrollregion=(0, 0, new_width, new_height))

    # Прокручиваем так, чтобы под курсором осталась та же точка
    canvas2.xview_moveto((mouse_x * zoom_factor - event.x) / new_width)
    canvas2.yview_moveto((mouse_y * zoom_factor - event.y) / new_height)

    # где у тебя есть текущее значение масштаба, например scale или canvas_scale:
    try:
        status_set(zoom=f"{int(canvas_scale * 100)}%")
    except Exception:
        pass

def update_coordinates(event):
    global x_start, y_start, x_end, y_end, coordinates_entry
    try:
        # Получаем текст из поля и разбиваем его на координаты
        coordinates = coordinates_entry.get().split(",")

        # Проверяем, что получено 4 значения
        if len(coordinates) == 4:
            x_start, y_start, x_end, y_end = map(int, coordinates)
            logger.info(f"Обновленные координаты: x_start={x_start}, y_start={y_start}, x_end={x_end}, y_end={y_end}")
            draw_selection()
        else:
            logger.info("Ошибка: Введите 4 координаты, разделенные запятыми")
    except ValueError:
        # Игнорируем ошибку, если ввод некорректен
        logger.info("Ошибка: Неверный формат координат")


# Процедура записи параметров в файл .env
def save_env(x_start, y_start, x_end, y_end, regex_pattern):
    set_key(ENV_FILE, "X_START", str(x_start))
    set_key(ENV_FILE, "Y_START", str(y_start))
    set_key(ENV_FILE, "X_END", str(x_end))
    set_key(ENV_FILE, "Y_END", str(y_end))
    set_key(ENV_FILE, "REGEX_PATTERN", str(regex_pattern))
    # regex_pattern = regex_pattern_entry.get()  # Получение текущего шаблона


# Функция для чтения параметров из .env файла
def read_env():
    global x_start, y_start, x_end, y_end, regex_pattern, selected_areas

    load_dotenv(ENV_FILE)
    try:
        x_start = int(os.getenv("X_START", DEFAULT_COORDINATES2["X_START"]))
        y_start = int(os.getenv("Y_START", DEFAULT_COORDINATES2["Y_START"]))
        x_end = int(os.getenv("X_END", DEFAULT_COORDINATES2["X_END"]))
        y_end = int(os.getenv("Y_END", DEFAULT_COORDINATES2["Y_END"]))
        regex_pattern = os.getenv("REGEX_PATTERN", DEFAULT_COORDINATES2["REGEX_PATTERN"])

        # Обновляем selected_areas
        selected_areas = [(None, x_start, y_start, x_end, y_end)]

        # Обновляем поле ввода координат
        if "coordinates_entry" in globals():
            coordinates_entry.delete(0, tk.END)
            coordinates_entry.insert(0, f"{x_start},{y_start},{x_end},{y_end}")

        logger.info(f"Загружены координаты из .env: x={x_start}, y={y_start}, w={x_end - x_start}, h={y_end - y_start}")

    except Exception as e:
        logger.info(f"[ERROR] Ошибка при загрузке координат из .env: {e}")
        # Устанавливаем значения по умолчанию
        # x_start, y_start, x_end, y_end = (
        #     DEFAULT_COORDINATES2["x_start"],
        #     DEFAULT_COORDINATES2["y_start"],
        #     DEFAULT_COORDINATES2["x_end"],
        #     DEFAULT_COORDINATES2["y_end"],
        # )
        x_start, y_start, x_end, y_end = (
            DEFAULT_COORDINATES2["X_START"],
            DEFAULT_COORDINATES2["Y_START"],
            DEFAULT_COORDINATES2["X_END"],
            DEFAULT_COORDINATES2["Y_END"],
        )

        selected_areas = [(None, x_start, y_start, x_end, y_end)]


# Обработчик выхода
def on_closing():
    global root, x_start, y_start, x_end, y_end, regex_pattern
    if root is not None:  # Проверяем, что окно еще существует
        save_env(x_start, y_start, x_end, y_end, regex_pattern)
        root.destroy()  # Закрыть главное окно



class TextRedirector:
    def __init__(self, widget):
        self.widget = widget

    def write(self, text):
        self.widget.insert(tk.END, text)
        self.widget.see(tk.END)  # Прокрутка текста вниз

    def flush(self):
        pass  # Для совместимости с sys.stdout


def _get_match_and_recognized(values):
    """Возвращает (match, recognized) с учётом возможной вставки колонки invoice."""
    # Старый порядок: [number, expected, recognized, match, score]
    # Новый порядок:  [number, expected, invoice, recognized, match, score]
    match = (values[4] if len(values) > 4 else (values[3] if len(values) > 3 else "")) or ""
    recognized_in_row = (values[3] if len(values) > 3 else (values[2] if len(values) > 2 else "")) or ""
    return match.strip(), recognized_in_row.strip()

def _find_invoice_by_expected(expected_value):
    """Находит номер накладной по совпадению expected == expected_value.
    Сначала берём из entry['xls_id'], при наличии колонки invoice — можно вытащить и из values[2].
    """
    if not expected_value:
        return ""
    ev = expected_value.strip()
    for e in table_entries:
        vals = list(tree.item(e["item_id"], "values"))
        if len(vals) > 1 and (vals[1] or "").strip() == ev:
            # приоритет — из структуры entry
            inv = (e.get("xls_id") or "").strip()
            if not inv and len(vals) > 2:
                inv = (vals[2] or "").strip()  # если появится явная колонка invoice
            return inv
    return ""


@safe_execute
def save_current_page2():
    """Сохранение текущей страницы как изображения."""
    global current_page, pdf_path
    if not pdf_path:
        messagebox.showerror("Ошибка", "Не выбран PDF файл.")
        return

    try:
        with fitz.open(pdf_path) as pdf:
            if current_page < 0 or current_page >= pdf.page_count:
                messagebox.showerror("Ошибка", "Неверный номер страницы.")
                return

            page = pdf.load_page(current_page)
            pix = page.get_pixmap(dpi=200)
            page_image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

            # Берем имя из распознанного или совпадения
            recognized = (table_entries[current_page].get("recognized") or "").strip()
            match = (tree.item(table_entries[current_page]["item_id"], "values")[4] or "").strip()
            filename = match or recognized or f"page_{current_page + 1}"

            # ← ДОБАВИЛИ префикс из колонки 3 Excel:
            prefix = (table_entries[current_page].get("xls_id") or "").strip()
            if prefix:
                filename = f"{prefix}_{filename}"

            output_dir = Path(entry_pdf_path.get()).parent
            output_file = Path(output_dir) / f"{current_page + 1}_{filename}"
            page_image.save(f"{output_file}_full.jpg")

            if debug_mode.get() and hasattr(canvas2, "image") and canvas2.image:
                cropped_image.save(f"{output_file}_cropped.jpg")

            messagebox.showinfo("Успех", f"Страница сохранена как {output_file}_full.jpg")

    except Exception as e:
        logger.error(f"Ошибка при сохранении страницы: {e}", exc_info=True)

def save_current_page():
    """Сохранение текущей страницы как изображения.
    Базовое имя: match → recognized → page_N.
    Префикс накладной: ищем строку, где expected == (match|recognized), и берём её invoice (xls_id).
    """
    global current_page, pdf_path, table_entries, tree, debug_mode

    if not pdf_path:
        messagebox.showerror("Ошибка", "Не выбран PDF файл.")
        return

    try:
        with fitz.open(pdf_path) as pdf:
            if current_page < 0 or current_page >= pdf.page_count:
                messagebox.showerror("Ошибка", "Неверный номер страницы.")
                return

            page = pdf.load_page(current_page)
            pix = page.get_pixmap(dpi=200)
            page_image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

            # --- 1) Соберём данные строки и базовое имя -----------------------
            row_values = list(tree.item(table_entries[current_page]["item_id"], "values"))

            # match и recognized по новой схеме (с учётом вставленной колонки invoice)
            match_cell = (row_values[4] if len(row_values) > 4 and row_values[4] else "").strip()
            # recognized берём сперва из entry (если уже сохранён там), затем из таблицы
            recognized_cell_in_table = (row_values[3] if len(row_values) > 3 and row_values[3] else "").strip()
            recognized = (table_entries[current_page].get("recognized") or recognized_cell_in_table or "").strip()

            # Базовое имя: match → recognized → page_N
            basename = match_cell or recognized or f"page_{current_page + 1}"

            # --- 2) Найдём накладную по expected == (match|recognized) --------
            invoice_prefix = ""
            search_key = (match_cell or recognized).strip()
            if search_key:
                for e in table_entries:
                    vals = list(tree.item(e["item_id"], "values"))
                    expected_val = (vals[1] if len(vals) > 1 and vals[1] else "").strip()
                    if expected_val == search_key:
                        # приоритетно берём из структуры записи, иначе из видимой таблицы (колонка invoice)
                        invoice_prefix = (e.get("xls_id") or (vals[2] if len(vals) > 2 else "") or "").strip()
                        if invoice_prefix:
                            break

            # --- 3) Итоговое имя файла ---------------------------------------
            filename = f"{invoice_prefix}_{basename}" if invoice_prefix else basename

            # Папка рядом с исходным PDF
            output_dir = Path(entry_pdf_path.get()).parent
            output_file_root = Path(output_dir) / f"{current_page + 1}_{filename}"

            # Сохранить полную страницу
            page_image.save(f"{output_file_root}_full.jpg")

            # При Debug — сохранить кроп, если он есть
            if debug_mode.get():
                # Пытаемся использовать глобальный cropped_image, если он есть
                cropped = globals().get("cropped_image", None)
                if cropped is not None:
                    try:
                        cropped.save(f"{output_file_root}_cropped.jpg")
                    except Exception:
                        pass
                # Либо canvas2.image, если сохранён там
                elif 'canvas2' in globals() and hasattr(canvas2, "image") and canvas2.image is not None:
                    try:
                        canvas2.image.save(f"{output_file_root}_cropped.jpg")
                    except Exception:
                        pass

            messagebox.showinfo("Успех", f"Страница сохранена как {output_file_root}_full.jpg")

    except Exception as e:
        logger.error(f"Ошибка при сохранении страницы: {e}", exc_info=True)
        messagebox.showerror("Ошибка", f"Не удалось сохранить страницу: {e}")



def toggle_extra_options():
    global frame_extra, extra_mode
    if extra_mode.get():
        if not frame_extra.winfo_ismapped():
            # frame_extra.pack(fill="x", padx=10, pady=5, before=frame_main)
            frame_extra.pack(fill="x", padx=10, pady=5, before=frame_canvases)
    else:
        if frame_extra.winfo_ismapped():
            frame_extra.pack_forget()

def on_link_click(event, url):
    webbrowser.open(url)  # Открывает ссылку в браузере

def insert_link(text_widget, text, url):
    text_widget.insert(tk.INSERT, text, "hyperlink")
    text_widget.tag_bind("hyperlink", "<Button-1>", lambda e: on_link_click(e, url))  # Привязка клика
    text_widget.insert(tk.INSERT, "\n")  # Переход на новую строку



# def create_interface2():
#     global root, entry_pdf_path, canvas, label_page_number, label_page_size, label_scale, coordinates_entry
#     global canvas2, canvas2_scale, label_coordinates, text_output, regex_pattern_entry, recognition_mode
#     global ocr_engine_var, table_frame, tree
#     global selected_areas
#     global debug_mode
#     global extra_mode, frame_extra
#     global frame_main, frame_canvases
#
#     root = tk.Tk()
#     apply_minimal_theme(root)
#
#     extra_mode = tk.BooleanVar(value=False)
#     root.title(f"Распознавание текста из PDF - Текущая версия программы: {__version__}")
#     root.protocol("WM_DELETE_WINDOW", on_closing)
#
#     window_width = 1400
#     window_height = 800
#     screen_width = root.winfo_screenwidth()
#     screen_height = root.winfo_screenheight()
#     x = (screen_width // 2) - (window_width // 2)
#     y = (screen_height // 2) - (window_height // 2)
#     root.geometry(f"{window_width}x{window_height}+{x}+{y}")
#
#     frame_top = tk.Frame(root)
#     # frame_top.pack(fill="x", expand=True)
#     frame_top.pack(side=tk.TOP, fill="x", expand=False, anchor="n")
#     frame_top.columnconfigure(2, weight=1)
#
#     frame_extra = tk.Frame(root)
#     if extra_mode.get():
#         frame_extra.pack(pady=5, padx=10, fill="x")
#
#     frame_main = tk.Frame(root)
#     frame_main.pack(pady=5, padx=10, fill="x")
#
#     button_style = {"width": 20, "anchor": "center"}
#
#     btn_select_pdf = tk.Button(frame_top, text="Выбрать PDF", command=select_pdf, **button_style)
#     btn_load_registry = tk.Button(frame_top, text="Выбрать XLS", command=load_registry, **button_style)
#     entry_pdf_path = tk.Entry(frame_top)
#     btn_recognize = tk.Button(frame_top, text="Запуск распознавания", command=start_recognition_thread, **button_style)
#     btn_match = tk.Button(frame_top, text="Сопоставить", command=match_with_expected, **button_style)
#     btn_save = tk.Button(frame_top, text="Сохранить результаты", command=lambda: save_results(btn_save), **button_style)
#
#     btn_select_pdf.grid(row=0, column=0, padx=5, pady=5)
#     btn_load_registry.grid(row=0, column=1, padx=5, pady=5)
#     entry_pdf_path.grid(row=0, column=2, padx=5, pady=5, sticky="we")
#     btn_recognize.grid(row=0, column=3, padx=5, pady=5)
#     btn_match.grid(row=0, column=4, padx=5, pady=5)
#     btn_save.grid(row=0, column=5, padx=5, pady=5)
#
#     frame_left = tk.Frame(frame_main)
#     frame_left.pack(side=tk.LEFT, fill="x", expand=False)
#
#     button_width = 15
#     btn_prev = tk.Button(frame_left, text="← Назад", command=prev_page, width=button_width)
#     btn_next = tk.Button(frame_left, text="Вперед →", command=next_page, width=button_width)
#     btn_check = tk.Button(frame_left, text="Проверить лист", command=check_image, width=button_width)
#     btn_save_page = tk.Button(frame_left, text="Сохранить лист", command=save_current_page, width=button_width)
#     extra_checkbutton = tk.Checkbutton(frame_left, text="Options", variable=extra_mode, command=toggle_extra_options)
#
#     btn_prev.pack(side=tk.LEFT, padx=2)
#     btn_next.pack(side=tk.LEFT, padx=2)
#     btn_check.pack(side=tk.LEFT, padx=2)
#     btn_save_page.pack(side=tk.LEFT, padx=2)
#     extra_checkbutton.pack(side=tk.LEFT, padx=2)
#
#     separator1 = ttk.Separator(frame_main, orient="vertical")
#     separator1.pack(side=tk.LEFT, fill="y", padx=5)
#
#     frame_left_extra = tk.Frame(frame_extra)
#     frame_left_extra.pack(side=tk.LEFT, fill="x", expand=False)
#
#     recognition_mode = tk.IntVar(value=0)
#     adv_checkbutton = tk.Checkbutton(frame_left_extra, text="Advance", variable=recognition_mode)
#
#     debug_mode = tk.BooleanVar(value=False)
#     debug_checkbutton = tk.Checkbutton(frame_left_extra, text="Debug", variable=debug_mode, command=update_debug_mode)
#
#     adv_checkbutton.pack(side=tk.LEFT, padx=2)
#     debug_checkbutton.pack(side=tk.LEFT, padx=2)
#
#     # btn_theme = ttk.Button(frame_left_extra, text="Тема", command=toggle_theme)
#     # btn_theme.grid(row=0, column=6, padx=6, pady=8, sticky="e")
#
#
#     separator2 = ttk.Separator(frame_extra, orient="vertical")
#     separator2.pack(side=tk.LEFT, fill="y", padx=5)
#
#     frame_center = tk.Frame(frame_extra)
#     frame_center.pack(side=tk.LEFT, fill="x", expand=True)
#     ocr_container = tk.Frame(frame_center)
#     ocr_container.pack(fill="x", expand=True)
#
#     lbl_ocr = tk.Label(ocr_container, text="OCR движок:")
#     ocr_engine_var = tk.StringVar(value="Tesseract")
#     ocr_options = ["Tesseract","EasyOCR","PaddleOCR"]
#     # if EASYOCR_AVAILABLE:
#     #     ocr_options.append("EasyOCR")
#     # if PADDLEOCR_AVAILABLE:
#     #     ocr_options.append("PaddleOCR")
#     ocr_menu = tk.OptionMenu(ocr_container, ocr_engine_var, *ocr_options)
#     btn_init_ocr = tk.Button(ocr_container, text="Инициализировать", command=init_ocr_engine)
#
#     lbl_ocr.pack(side=tk.LEFT, padx=2)
#     ocr_menu.pack(side=tk.LEFT, padx=2)
#     btn_init_ocr.pack(side=tk.LEFT, padx=2)
#
#     separator3 = ttk.Separator(frame_extra, orient="vertical")
#     separator3.pack(side=tk.LEFT, fill="y", padx=5)
#
#     frame_right = tk.Frame(frame_extra)
#     frame_right.pack(side=tk.RIGHT, fill="x", expand=False)
#
#     lbl_pattern = tk.Label(frame_right, text="Шаблон:")
#     regex_pattern_entry = tk.Entry(frame_right, width=25)
#     coordinates_entry = tk.Entry(frame_right, width=20)
#
#     lbl_pattern.pack(side=tk.LEFT, padx=2)
#     regex_pattern_entry.pack(side=tk.LEFT, padx=2)
#     coordinates_entry.pack(side=tk.LEFT, padx=2)
#
#     regex_pattern_entry.insert(0, regex_pattern)
#     coordinates_entry.bind("<KeyRelease>", update_coordinates)
#
#     # Создание холстов и таблицы
#     frame_canvases = tk.Frame(root)
#     frame_canvases.pack(pady=10, padx=10, fill="both", expand=True)
#
#     canvas_width = 750 // 2
#     canvas_height = 500
#
#     # Левый холст (PDF)
#     canvas = tk.Canvas(frame_canvases, width=canvas_width, height=canvas_height, bg="grey")
#     canvas.pack(side=tk.LEFT, anchor=tk.N, padx=5, pady=10)
#
#     # Правый холст (выделенная область)
#     canvas2 = tk.Canvas(frame_canvases, width=canvas_width, height=canvas_height, bg="grey")
#     canvas2.pack(side=tk.LEFT, anchor=tk.N, padx=5, pady=10)
#
#     # Обёртка для таблицы Treeview
#     table_frame = tk.Frame(frame_canvases)
#     table_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
#
#     # Создание Treeview с прокруткой
#     tree_scroll = tk.Scrollbar(table_frame)
#     tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
#
#     tree = ttk.Treeview(table_frame, yscrollcommand=tree_scroll.set, selectmode="browse")
#     tree.pack(fill=tk.BOTH, expand=True)
#
#     tree_scroll.config(command=tree.yview)
#
#     # Настройка колонок
#     tree["columns"] = ("number", "expected", "recognized", "match", "score")
#     tree.column("#0", width=0, stretch=tk.NO)  # Скрытая колонка
#     tree.column("number", width=50, anchor=tk.CENTER)
#     tree.column("expected", width=150, anchor=tk.W)
#     # tree.column("expected", width=0, stretch=tk.NO)
#     tree.column("recognized", width=150, anchor=tk.W)
#     tree.column("match", width=150, anchor=tk.W)
#     tree.column("score", width=50, anchor=tk.CENTER)
#
#     # Заголовки
#     tree.heading("number", text="№")
#     tree.heading("expected", text="Контейнер из XLS")
#     tree.heading("recognized", text="Контейнер распознанный")
#     tree.heading("match", text="Совпадение")
#     tree.heading("score", text="Коэффициент")
#
#     # Привязка события двойного клика для перехода к странице
#     tree.bind("<Button-1>", on_tree_click)
#     tree.bind("<Return>", on_tree_enter)
#
#     # Настройка обработчиков событий для холстов
#     canvas.bind("<Button-1>", define_coordinates)
#     canvas.bind("<B1-Motion>", draw_rectangle)
#     canvas.bind("<ButtonRelease-1>", finish_coordinates)
#     canvas.bind("<MouseWheel>", zoom_canvas)
#     canvas2.bind("<MouseWheel>", zoom_canvas2)
#
#     # Текстовое поле вывода
#     text_output = scrolledtext.ScrolledText(root, width=100, height=10)
#     text_output.pack(side=tk.BOTTOM, fill="x", pady=10, padx=10)
#     text_output.config(state="normal")
#     sys.stdout = TextRedirector(text_output)
#
#     # Добавляем гиперссылки
#     text_output.tag_configure("hyperlink", foreground="blue", underline=True)
#
#     set_default_coordinates(coordinates_entry)
#
#     toggle_extra_options()
#     logger.update_gui_handler(text_output)

def create_interface():
    global root, entry_pdf_path, canvas, label_page_number, label_page_size, label_scale, coordinates_entry
    global canvas2, canvas2_scale, label_coordinates, text_output, regex_pattern_entry, recognition_mode
    global ocr_engine_var, table_frame, tree
    global selected_areas
    global debug_mode
    global extra_mode, frame_extra
    global frame_main, frame_canvases, current_theme

    # --- окно + тема ---
    root = tk.Tk()

    print("[DEBUG] Тема при старте:", current_theme)

    extra_mode = tk.BooleanVar(value=False)
    root.title(f"Распознавание текста из PDF - Текущая версия программы: {__version__}")
    root.protocol("WM_DELETE_WINDOW", on_closing)

    # Центр окна
    window_width, window_height = 1400, 800
    root.update_idletasks()
    sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
    x = (sw - window_width) // 2
    y = (sh - window_height) // 2
    root.geometry(f"{window_width}x{window_height}+{x}+{y}")

    # --- Верхняя панель (прилипает к верху) ---
    frame_top = ttk.Frame(root, style="Toolbar.TFrame")
    frame_top.pack(side=tk.TOP, fill="x")
    frame_top.grid_columnconfigure(2, weight=1)

    btn_select_pdf    = ttk.Button(frame_top, text="Выбрать PDF",         command=select_pdf)
    btn_load_registry = ttk.Button(frame_top, text="Выбрать XLS",         command=load_registry)
    entry_pdf_path    = ttk.Entry(frame_top)
    btn_recognize     = ttk.Button(frame_top, text="Запуск распознавания", command=start_recognition_thread)
    btn_match         = ttk.Button(frame_top, text="Сопоставить",         command=match_with_expected)
    btn_save          = ttk.Button(frame_top, text="Сохранить результаты", command=lambda: save_results(btn_save))

    btn_select_pdf.grid(   row=0, column=0, padx=6, pady=8, sticky="w")
    btn_load_registry.grid(row=0, column=1, padx=6, pady=8, sticky="w")
    entry_pdf_path.grid(   row=0, column=2, padx=6, pady=8, sticky="we")
    btn_recognize.grid(   row=0, column=3, padx=6, pady=8, sticky="e")
    btn_match.grid(       row=0, column=4, padx=6, pady=8, sticky="e")
    btn_save.grid(        row=0, column=5, padx=6, pady=8, sticky="e")

    ttk.Separator(root, orient="horizontal").pack(side=tk.TOP, fill="x")

    # --- Основная панель навигации ---
    frame_main = ttk.Frame(root)
    frame_main.pack(pady=6, padx=10, fill="x")

    frame_left = ttk.Frame(frame_main)
    frame_left.pack(side=tk.LEFT, fill="x", expand=False)

    button_width = 16
    btn_prev      = ttk.Button(frame_left, text="← Назад",        command=prev_page,         width=button_width)
    btn_next      = ttk.Button(frame_left, text="Вперед →",       command=next_page,         width=button_width)
    btn_check     = ttk.Button(frame_left, text="Проверить лист", command=check_image,       width=button_width)
    btn_save_page = ttk.Button(frame_left, text="Сохранить лист", command=save_current_page, width=button_width)

    btn_prev.pack(side=tk.LEFT, padx=4, pady=2)
    btn_next.pack(side=tk.LEFT, padx=4, pady=2)
    btn_check.pack(side=tk.LEFT, padx=4, pady=2)
    btn_save_page.pack(side=tk.LEFT, padx=4, pady=2)

    # ← ВЕРНУЛИ кнопку Options (переключает панель frame_extra)
    ttk.Separator(frame_main, orient="vertical").pack(side=tk.LEFT, fill="y", padx=8)
    options_btn = ttk.Checkbutton(frame_left, text="Options", variable=extra_mode, command=toggle_extra_options)
    options_btn.pack(side=tk.LEFT, padx=6)

    # --- Панель дополнительных опций (скрываемая) ---
    frame_extra = ttk.Frame(root)
    if extra_mode.get():
        frame_extra.pack(pady=4, padx=10, fill="x")

    # ВНИМАНИЕ: frame_left_extra снова ВНУТРИ frame_extra (как у тебя было)
    frame_left_extra = ttk.Frame(frame_extra)
    frame_left_extra.pack(side=tk.LEFT, fill="x", expand=False)

    recognition_mode = tk.IntVar(value=0)
    debug_mode = tk.BooleanVar(value=False)

    adv_checkbutton   = ttk.Checkbutton(frame_left_extra, text="Advance", variable=recognition_mode)
    debug_checkbutton = ttk.Checkbutton(frame_left_extra, text="Debug",   variable=debug_mode, command=update_debug_mode)
    adv_checkbutton.pack(side=tk.LEFT, padx=4)
    debug_checkbutton.pack(side=tk.LEFT, padx=4)
    btn_theme = ttk.Button(frame_left_extra, text="Тема", command=toggle_theme)
    btn_theme.pack(side=tk.LEFT, padx=6)


    ttk.Separator(frame_extra, orient="vertical").pack(side=tk.LEFT, fill="y", padx=8)

    # Центр блока доп. опций: OCR инициализация
    frame_center = ttk.Frame(frame_extra)
    frame_center.pack(side=tk.LEFT, fill="x", expand=True)

    ocr_container = ttk.Frame(frame_center)
    ocr_container.pack(fill="x", expand=True)

    ttk.Label(ocr_container, text="OCR движок:").pack(side=tk.LEFT, padx=4)
    ocr_engine_var = tk.StringVar(value="Tesseract")
    ocr_options = ["Tesseract", "EasyOCR", "PaddleOCR"]
    ocr_menu = ttk.Combobox(ocr_container, textvariable=ocr_engine_var, values=ocr_options, state="readonly", width=16)
    ocr_menu.pack(side=tk.LEFT, padx=4)
    ttk.Button(ocr_container, text="Инициализировать", command=init_ocr_engine).pack(side=tk.LEFT, padx=4)

    ttk.Separator(frame_extra, orient="vertical").pack(side=tk.LEFT, fill="y", padx=8)

    # Правый блок: regex и координаты
    frame_right = ttk.Frame(frame_extra)
    frame_right.pack(side=tk.RIGHT, fill="x", expand=False)

    ttk.Label(frame_right, text="Шаблон:").pack(side=tk.LEFT, padx=4)
    regex_pattern_entry = ttk.Entry(frame_right, width=28)
    regex_pattern_entry.pack(side=tk.LEFT, padx=4)

    coordinates_entry = ttk.Entry(frame_right, width=22)
    coordinates_entry.pack(side=tk.LEFT, padx=4)

    regex_pattern_entry.insert(0, regex_pattern)
    coordinates_entry.bind("<KeyRelease>", update_coordinates)

    # --- Центральная область: 2 Canvas + таблица ---
    frame_canvases = ttk.Frame(root)
    frame_canvases.pack(pady=10, padx=10, fill="both", expand=True)

    canvas_width = 750 // 2
    canvas_height = 500

    canvas = tk.Canvas(frame_canvases, width=canvas_width, height=canvas_height, bg="#F3F4F6", highlightthickness=0)
    canvas.pack(side=tk.LEFT, anchor=tk.N, padx=6, pady=10)

    canvas2 = tk.Canvas(frame_canvases, width=canvas_width, height=canvas_height, bg="#F3F4F6", highlightthickness=0)
    canvas2.pack(side=tk.LEFT, anchor=tk.N, padx=6, pady=10)

    ttk.Separator(frame_canvases, orient="vertical").pack(side=tk.LEFT, fill="y", padx=8)

    table_frame = ttk.Frame(frame_canvases)
    table_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    tree_scroll = ttk.Scrollbar(table_frame, orient="vertical")
    tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

    tree = ttk.Treeview(table_frame, yscrollcommand=tree_scroll.set, selectmode="browse")
    tree.pack(fill=tk.BOTH, expand=True)
    tree_scroll.config(command=tree.yview)

    tree["columns"] = ("number", "expected", "invoice", "recognized", "match", "score")
    tree.column("#0", width=0, stretch=tk.NO)
    tree.column("number",     width=60,  anchor=tk.CENTER)
    tree.column("expected",   width=0,   minwidth=0, stretch=tk.NO)
    tree.column("invoice",    width=0,   stretch=tk.NO)  # скрыто по умолчанию
    tree.column("recognized", width=170, anchor=tk.W)
    tree.column("match",      width=170, anchor=tk.W)
    tree.column("score",      width=80,  anchor=tk.CENTER)

    tree.heading("number", text="№")
    tree.heading("expected", text="Контейнер из XLS")
    tree.heading("invoice", text="Накладная (XLS)")
    tree.heading("recognized", text="Контейнер распознанный")
    tree.heading("match", text="Совпадение")
    tree.heading("score", text="Коэффициент")

    # «Зебра» для строк (после заполнения можно вызвать ещё раз)
    try:
        root._style_helpers["style_treeview_stripes"](tree)
    except Exception:
        pass

    # Бинды
    tree.bind("<Button-1>", on_tree_click)
    tree.bind("<Return>", on_tree_enter)

    canvas.bind("<Button-1>", define_coordinates)
    canvas.bind("<B1-Motion>", draw_rectangle)
    canvas.bind("<ButtonRelease-1>", finish_coordinates)
    canvas.bind("<MouseWheel>", zoom_canvas)
    canvas2.bind("<MouseWheel>", zoom_canvas2)

    # --- Лог-панель ---
    text_output = scrolledtext.ScrolledText(root, height=8)
    text_output.pack(side=tk.BOTTOM, fill="x", pady=8, padx=10)
    text_output.config(state="normal")
    sys.stdout = TextRedirector(text_output)
    text_output.tag_configure("hyperlink", foreground="blue", underline=True)

    # Значения по умолчанию и логгер
    set_default_coordinates(coordinates_entry)
    toggle_extra_options()   # учитывает текущее extra_mode и (un)pack'ает frame_extra
    logger.update_gui_handler(text_output)

    # --- Статус-бар снизу ---
    global status_page_var, status_zoom_var, status_size_var, status_msg_var
    status_page_var = tk.StringVar(value="Стр: —/—")
    status_zoom_var = tk.StringVar(value="Масштаб: 100%")
    status_size_var = tk.StringVar(value="Размер: —×—")
    status_msg_var = tk.StringVar(value="Готово")

    statusbar = ttk.Frame(root, style="Toolbar.TFrame")
    # Порядок важен: этот pack ставим ПОСЛЕДНИМ, чтобы статус-бар был внизу
    statusbar.pack(side=tk.BOTTOM, fill="x")

    ttk.Label(statusbar, textvariable=status_page_var).pack(side=tk.LEFT, padx=8, pady=4)
    ttk.Label(statusbar, text="|").pack(side=tk.LEFT, padx=6)
    ttk.Label(statusbar, textvariable=status_zoom_var).pack(side=tk.LEFT, padx=8)
    ttk.Label(statusbar, text="|").pack(side=tk.LEFT, padx=6)
    ttk.Label(statusbar, textvariable=status_size_var).pack(side=tk.LEFT, padx=8)

    # Сообщения статуса — справа
    ttk.Label(statusbar, textvariable=status_msg_var).pack(side=tk.RIGHT, padx=8)

    apply_minimal_theme(root, "light")


def update_debug_mode():
    """Обновляет режим отладки при изменении чекбокса"""
    global debug_mode, tree, text_output
    dbg = bool(debug_mode.get())
    logger.info(f"Debug mode: {dbg}")

    import logging
    if dbg:
        # Показать "Контейнер из XLS"
        tree.column("expected", width=170, minwidth=50, stretch=tk.YES)
        tree.heading("expected", text="Контейнер из XLS")
        tree.column("invoice", width=140, minwidth=50, stretch=tk.YES)
        tree.heading("invoice", text="Накладная из XLS")
        # Включить уровень DEBUG
        logger.logger.setLevel(logging.DEBUG)
        logger.info("Уровень логов переключен на DEBUG")
    else:
        # Скрыть "Контейнер из XLS"
        tree.column("expected", width=0, minwidth=0, stretch=tk.NO)
        tree.heading("expected", text="")
        tree.column("invoice", width=0, minwidth=0, stretch=tk.NO)
        tree.heading("invoice", text="")

        # Вернуть уровень INFO
        logger.logger.setLevel(logging.INFO)
        logger.info("Уровень логов переключен на INFO")

    # Обновить GUI-хендлер, чтобы он подтянул новый уровень
    try:
        logger.update_gui_handler(text_output)
    except Exception:
        pass




def on_tree_click(event):
    """Обработчик кликов по Treeview"""
    global last_click_time

    # Определяем элемент, по которому кликнули
    region = tree.identify_region(event.x, event.y)
    if region != "cell":
        return

    current_time = event.time
    is_double_click = (current_time - last_click_time) < DOUBLE_CLICK_DELAY
    last_click_time = current_time

    item = tree.identify_row(event.y)
    column = tree.identify_column(event.x)

    # Всегда выполняем переход к странице
    goto_page(item)

    # Если это двойной клик - дополнительно открываем редактор
    if is_double_click and column == "#5":  # Только для столбца "Совпадение"
        edit_cell(item, column)


def on_tree_enter(event):
    item = tree.focus() or (tree.selection()[0] if tree.selection() else None)
    if not item:
        children = tree.get_children()
        if not children:
            return "break"
        item = children[0]
        tree.selection_set(item)
        tree.focus(item)

    tree.see(item)
    goto_page(item)           # как при клике
    edit_cell(item, "#5")     # открыть редактор "Совпадение"
    return "break"



def goto_page(item):
    """Переход к указанной странице"""
    global current_page
    values = tree.item(item, "values")
    if values and values[0]:  # values[0] - номер страницы
        try:
            current_page = int(values[0]) - 1
            load_page()
        except ValueError:
            pass
    # после смены current_page и отрисовки страницы:
    try:
        total = pdf_doc.page_count
        status_set(page=current_page + 1, total=total, msg="Готово")
    except Exception:
        pass



def edit_cell(item, column):
    global expected_containers  # предполагаем, что это список контейнеров из XLS

    # Обновляем список контейнеров для автоподсказки — только при несовпадении
    expected_containers = []
    for child in tree.get_children():
        values = tree.item(child, "values")
        try:
            container = values[1]  # "Контейнер из XLS"
            # score = float(values[4]) if values[4] else 0
            # if container and score < 1.0:
            #     expected_containers.append(container)
            expected_containers.append(container)
        except (IndexError, ValueError):
            continue

    # Получаем координаты и текущее значение
    x, y, width, height = tree.bbox(item, column)
    current_value = tree.item(item, "values")[4]

    # Создаём поле ввода
    first_input = {"done": False}  # ← флаг для сброса текста при первом вводе
    entry_edit = tk.Entry(tree, borderwidth=0, font=("Arial", 10))
    entry_edit.place(x=x, y=y, width=width, height=height, anchor=tk.NW)
    entry_edit.insert(0, current_value)
    entry_edit.focus_set()

    # Создаём Listbox для автоподсказок
    listbox = tk.Listbox(tree, height=5)
    listbox.place(x=x, y=y + height, width=width)

    def update_listbox():
        typed = entry_edit.get()
        # matches = [c for c in expected_containers if c.startswith(typed)]
        matches = [c for c in expected_containers if typed in c]
        listbox.delete(0, tk.END)
        for match in matches:
            listbox.insert(tk.END, match)
        if matches:
            listbox.place(x=x, y=y + height, width=width)
        else:
            listbox.place_forget()

    def on_key_release(event):
        if not first_input["done"]:
            if event.char and len(event.char) == 1:
                char = event.char.upper()
                entry_edit.delete(0, tk.END)
                entry_edit.insert(0, char)
                entry_edit.icursor(1)
                first_input["done"] = True
                update_listbox()
            return

        # if not first_input["done"]:
        #     char = event.char.upper()  # сохраняем первый введённый символ (в верхнем регистре)
        #     entry_edit.delete(0, tk.END)
        #     entry_edit.insert(0, char)
        #     entry_edit.icursor(1)
        #     first_input["done"] = True
        #     update_listbox()
        #     return

        if event.keysym in ("Up", "Down", "Return"):
            return

        # Преобразуем весь текст в верхний регистр
        pos = entry_edit.index(tk.INSERT)
        text = entry_edit.get().upper()
        entry_edit.delete(0, tk.END)
        entry_edit.insert(0, text)
        entry_edit.icursor(pos)

        update_listbox()

    def on_entry_key2(event):
        if event.keysym == "Down" and listbox.size() > 0:
            listbox.focus_set()
            listbox.selection_clear(0, tk.END)
            listbox.selection_set(0)
            listbox.activate(0)
            return "break"

        # elif event.keysym == "Return":
        #     if listbox.winfo_ismapped():
        #         if listbox.size() == 1:
        elif event.keysym == "Return" and listbox.winfo_ismapped() and listbox.size() == 1:
            # Автоматически подставляем единственный вариант
            selected = listbox.get(0)
            entry_edit.delete(0, tk.END)
            entry_edit.insert(0, selected)
            listbox.place_forget()
            entry_edit.focus_set()
                # иначе: сохраняем как есть
        save_edit()
        return "break"

    def on_entry_key3(event):
        # Стрелка вниз: перейти в список подсказок
        if event.keysym == "Down" and listbox.size() > 0:
            listbox.focus_set()
            listbox.selection_clear(0, tk.END)
            listbox.selection_set(0)
            listbox.activate(0)
            return "break"

        # Enter: принять единственную подсказку (если она одна) и сохранить
        if event.keysym == "Return":
            if listbox.winfo_ismapped() and listbox.size() == 1:
                selected = listbox.get(0)
                entry_edit.delete(0, tk.END)
                entry_edit.insert(0, selected)
                listbox.place_forget()
                entry_edit.focus_set()
            save_edit()  # ← сохраняем ТОЛЬКО на Enter
            return "break"

        # Esc: закрыть без сохранения (по желанию можно вернуть старое значение)
        if event.keysym == "Escape":
            listbox.place_forget()
            entry_edit.destroy()
            return "break"

        # Для других клавиш — ничего не делаем здесь.
        # Верхний регистр и фильтрацию ведём в on_key_release.
        return None

    def on_entry_key(event):
        # Стрелка вниз — уходим в список подсказок
        if event.keysym == "Down" and listbox.size() > 0:
            listbox.focus_set()
            listbox.selection_clear(0, tk.END)
            listbox.selection_set(0)
            listbox.activate(0)
            return "break"

        # Enter — сохранить и вернуть фокус в таблицу
        if event.keysym == "Return":
            if listbox.winfo_ismapped() and listbox.size() == 1:
                # автоподстановка единственного варианта
                selected = listbox.get(0)
                entry_edit.delete(0, tk.END)
                entry_edit.insert(0, selected)
                listbox.place_forget()
                entry_edit.focus_set()

            save_edit()
            # фокус и выделение строки в Treeview
            tree.focus(item)
            tree.selection_set(item)
            tree.see(item)
            tree.focus_set()
            return "break"

        # Esc — закрыть редактор без сохранения (и вернуть фокус в таблицу)
        if event.keysym == "Escape":
            listbox.place_forget()
            entry_edit.destroy()
            tree.focus_set()
            return "break"

        return None

    def on_listbox_key2(event):
        if event.keysym == "Return":
            on_listbox_select(None)
            return "break"
        elif event.keysym == "Escape":
            listbox.place_forget()
            entry_edit.focus_set()
            return "break"

    def on_listbox_key(event):
        if event.keysym == "Return":
            on_listbox_select(None)  # подставить выбранный вариант
            save_edit()
            tree.focus(item)
            tree.selection_set(item)
            tree.see(item)
            tree.focus_set()
            return "break"
        elif event.keysym == "Escape":
            listbox.place_forget()
            entry_edit.focus_set()
            return "break"

    def on_listbox_select(event):
        if listbox.curselection():
            selected = listbox.get(listbox.curselection())
            entry_edit.delete(0, tk.END)
            entry_edit.insert(0, selected)
            listbox.place_forget()
            entry_edit.focus_set()

    def save_edit(event=None):
        new_value = entry_edit.get()
        values = list(tree.item(item, "values"))
        values[4] = new_value

        # Обновим коэффициент
        recognized = values[3] if len(values) > 3 else ""
        if recognized and new_value:
            # score = is_similar_ratio(recognized, new_value)
            # values[4] = f"{score:.2f}"
            # update_row_color(item, score)
            tree.tag_configure("manual_edit", background="#ddaaff")
            tree.item(item, tags=("manual_edit",))

        tree.item(item, values=values)
        entry_edit.destroy()
        listbox.place_forget()

    def on_entry_focus_out(event):
        # Проверяем, ушёл ли фокус именно на listbox
        widget = event.widget.focus_get()
        if widget != listbox:
            save_edit()


    # Привязки
    entry_edit.bind("<KeyRelease>", on_key_release)
    entry_edit.bind("<FocusOut>", on_entry_focus_out)
    entry_edit.bind("<KeyPress>", on_entry_key)
    listbox.bind("<<ListboxSelect>>", on_listbox_select)
    listbox.bind("<Return>", on_listbox_key)
    listbox.bind("<Escape>", on_listbox_key)
    listbox.bind("<Double-Button-1>", on_listbox_select)
    listbox.bind("<FocusOut>", lambda e: listbox.place_forget())
    update_listbox()  # показать при открытии


def update_row_color(item, score):
    """Обновление цвета строки"""
    if score == 1.0:
        tree.tag_configure("exact", background="#a8e6a8")
        tree.item(item, tags=("exact",))
    elif score >= 0.7:
        tree.tag_configure("partial", background="#fff8a8")
        tree.item(item, tags=("partial",))
    else:
        tree.tag_configure("none", background="#ffaaaa")
        tree.item(item, tags=("none",))

def load_registry2():
    global table_entries, tree

    file_path = filedialog.askopenfilename(filetypes=[("Excel or CSV", "*.xlsx *.csv")])
    if not file_path:
        return

    records = []  # список кортежей (xls_id, container)

    try:
        suffix = Path(file_path).suffix.lower()

        if suffix == ".xlsx":
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            logger.info(f"Чтение данных из Excel файла: {file_path}")
            for row in sheet.iter_rows(min_row=5):
                # кол.3 → номер (xls_id), кол.4 → контейнер
                xls_id = ""
                container = ""
                if len(row) > 2 and row[2].value is not None:
                    xls_id = str(row[2].value).strip()
                if len(row) > 3:
                    cell = row[3].value
                    if cell is not None:
                        s = str(cell)
                        container = s.split("/")[-1].strip() if "/" in s else s.strip()
                records.append((xls_id, container))

        elif suffix == ".csv":
            with Path(file_path).open(encoding="utf-8") as f:
                reader = csv.reader(f)
                for idx, row in enumerate(reader):
                    if idx < 3:   # как и раньше — пропуск первых строк
                        continue
                    xls_id = row[2].strip() if len(row) > 2 and row[2] is not None else ""
                    cell = row[3] if len(row) > 3 else ""
                    container = cell.split("/")[-1].strip() if cell and "/" in cell else (cell.strip() if cell else "")
                    records.append((xls_id, container))

    except Exception as e:
        logger.error(f"Ошибка при загрузке реестра: {e}", exc_info=True)
        messagebox.showerror("Ошибка", f"Не удалось загрузить реестр: {e}")
        return

    # Обновляем таблицу: expected (values[1]) и служебное поле xls_id
    for i, (xls_id, code) in enumerate(records):
        if i < len(table_entries):
            table_entries[i]["code"] = code
            table_entries[i]["xls_id"] = xls_id

            current_values = list(tree.item(table_entries[i]["item_id"], 'values'))
            # гарантируем длину
            while len(current_values) < 5:
                current_values.append("")
            current_values[1] = code  # "Контейнер из XLS"
            tree.item(table_entries[i]["item_id"], values=current_values)

    # Возвращаем список контейнеров (как раньше) — если где-то используется
    return [c for _, c in records]

def load_registry():
    import csv
    import openpyxl
    from pathlib import Path
    from tkinter import filedialog, messagebox

    global table_entries, tree

    file_path = filedialog.askopenfilename(filetypes=[("Excel or CSV", "*.xlsx *.csv")])
    if not file_path:
        return

    records = []  # список кортежей (xls_id, container)

    try:
        suffix = Path(file_path).suffix.lower()

        if suffix == ".xlsx":
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            logger.info(f"Чтение данных из Excel файла: {file_path}")

            # Стартуем с 5-й строки: кол.3 → номер накладной (xls_id), кол.4 → контейнер
            for row in sheet.iter_rows(min_row=5):
                xls_id = ""
                container = ""

                # кол.3 (индекс 2)
                if len(row) > 2 and row[2].value is not None:
                    xls_id = str(row[2].value).strip()

                # кол.4 (индекс 3) — контейнер; если там путь, берём хвост после '/'
                if len(row) > 3:
                    cell = row[3].value
                    if cell is not None:
                        s = str(cell)
                        container = s.split("/")[-1].strip() if "/" in s else s.strip()

                records.append((xls_id, container))

        elif suffix == ".csv":
            logger.info(f"Чтение данных из CSV файла: {file_path}")
            with Path(file_path).open(encoding="utf-8") as f:
                reader = csv.reader(f)
                for idx, row in enumerate(reader):
                    # как и раньше — пропуск первых 3 строк (заголовки/служебные)
                    if idx < 3:
                        continue
                    xls_id = row[2].strip() if len(row) > 2 and row[2] is not None else ""
                    cell = row[3] if len(row) > 3 else ""
                    if cell:
                        cell = str(cell)
                        container = cell.split("/")[-1].strip() if "/" in cell else cell.strip()
                    else:
                        container = ""
                    records.append((xls_id, container))
        else:
            messagebox.showerror("Ошибка", f"Неподдерживаемый формат: {suffix}")
            return

    except Exception as e:
        logger.error(f"Ошибка при загрузке реестра: {e}", exc_info=True)
        messagebox.showerror("Ошибка", f"Не удалось загрузить реестр: {e}")
        return

    # Обновляем таблицу: expected (values[1]) и invoice (values[2]), плюс служебные поля в table_entries
    updated_rows = min(len(records), len(table_entries))
    for i in range(updated_rows):
        xls_id, code = records[i]

        # служебные поля
        table_entries[i]["code"] = code
        table_entries[i]["xls_id"] = xls_id

        # видимая таблица
        item_id = table_entries[i]["item_id"]
        current_values = list(tree.item(item_id, "values"))

        # гарантируем длину 6: [№, expected, invoice, recognized, match, score]
        while len(current_values) < 6:
            current_values.append("")

        current_values[1] = code    # Контейнер из XLS (expected)
        current_values[2] = xls_id  # Накладная (invoice)

        tree.item(item_id, values=tuple(current_values))

    if len(records) > len(table_entries):
        logger.warning(
            f"В реестре {len(records)} записей, в таблице {len(table_entries)} строк. "
            f"Лишние записи не отображены (нет соответствующих страниц PDF)."
        )

    # Возвращаем список контейнеров (как раньше), если где-то используется
    return [c for _, c in records]


def init_ocr_engine():
    global ocr_reader, EASYOCR_AVAILABLE, PADDLEOCR_AVAILABLE
    selected_engine = ocr_engine_var.get()

    try:
        if selected_engine == "EasyOCR":
            import easyocr

            ocr_reader = easyocr.Reader(["en"])  # Загрузка моделей
            EASYOCR_AVAILABLE = True
            messagebox.showinfo("Успех", "EasyOCR инициализирован!")

        elif selected_engine == "PaddleOCR":
            import paddle  # noqa
            from paddleocr import PaddleOCR

            ocr_reader = PaddleOCR(use_angle_cls=True, lang="en")
            PADDLEOCR_AVAILABLE = True
            messagebox.showinfo("Успех", "PaddleOCR инициализирован!")
        else:
            ocr_reader = None
            messagebox.showinfo("Инфо", "Используется Tesseract")

    except ImportError as e:
        error_msg = {"EasyOCR": "pip install easyocr", "PaddleOCR": "pip install paddlepaddle paddleocr"}.get(
            selected_engine, f"pip install {selected_engine.lower()}"
        )
        logger.error(f"Ошибка импорта: {e}\nНе хватает зависимостей!\nУстановите:\n{error_msg}")

    except Exception as e:
        logger.error("Ошибка", f"Не удалось инициализировать {selected_engine}: {e!s}")
        ocr_reader = None

def update_table(data, tree_widget):
    """Обновление таблицы Treeview"""
    global tree, table_entries, expected_containers

    # Очищаем таблицу
    for item in tree_widget.get_children():
        tree_widget.delete(item)

    table_entries = []
    expected_containers = [code for _, code in data]

    for number, code in data:
        # Добавляем строку в Treeview
        item_id = tree_widget.insert("", tk.END, values=(number, code, "", "", ""))

        # Сохраняем данные для дальнейшего использования
        entry = {
            "index": number,
            "code": code,
            "recognized": "",
            "item_id": item_id,  # Сохраняем ID элемента Treeview
        }
        table_entries.append(entry)

def on_match_edit(row_index):
    """Обработчик редактирования поля совпадения"""
    global table_entries
    if 0 <= row_index < len(table_entries):
        new_value = table_entries[row_index]["match_var"].get()
        # Здесь можно добавить логику обработки изменений
        logger.info(f"Изменено совпадение для строки {row_index + 1}: {new_value}")
        # Обновляем расчет коэффициента
        update_score(row_index)


def update_score(row_index):
    """Обновляет коэффициент совпадения"""
    global table_entries
    entry = table_entries[row_index]
    recognized = entry["recognized"]
    match = entry["match_var"].get()

    if recognized and match:
        score = is_similar_ratio(recognized, match)
        entry["label_score"].config(text=f"{score:.2f}")
        # Изменяем цвет в зависимости от результата
        bg_color = "lightgreen" if score >= 0.9 else "khaki" if score >= 0.7 else "tomato"
        entry["label_score"].config(bg=bg_color)

# Функция для проверки обновлений
def check_for_updates():
    threading.Thread(target=_check_updates, daemon=True).start()

def _check_updates():
    try:
        # URL для получения информации о релизах
        repo_url = "https://api.github.com/repos/vanitoo/pythonProject-OpenCV-PDF/releases/latest"
        response = requests.get(repo_url, timeout=5)

        if response.status_code == 200:
            latest_release = response.json()
            latest_version = latest_release["tag_name"].lstrip("v")  # Убираем префикс 'v'
            download_url = latest_release["assets"][0]["browser_download_url"]

            # Сравниваем текущую версию с последней на GitHub
            if compare_versions(__version__, latest_version):
                logger.info(f"Появилась новая версия {latest_version}, рекомендуется обновиться\n")
                # text_output.delete(1.0, tk.END)
                # text_output.insert(
                #     tk.END,
                #     f"Появилась новая версия {latest_version}, рекомендуется обновиться\n",
                # )
                logger.info(f"{download_url}")
                # text_output.insert(tk.END, download_url)
            else:
                logger.info("У вас последняя версия.\n")
                # text_output.delete(1.0, tk.END)
                # text_output.insert(tk.END, f"У вас последняя версия.\n")
    except Exception:
        pass  # Не блокировать интерфейс при ошибках

# Функция для сравнения версий
def compare_versions(current_version: str, latest_version: str) -> bool:
    try:
        current = tuple(map(int, current_version.split(".")))
        latest = tuple(map(int, latest_version.split(".")))
    except ValueError:
        raise ValueError("Version numbers must contain only integers separated by dots")

    # Compare component by component
    for current_part, latest_part in zip(current, latest, strict=True):
        if current_part < latest_part:
            return True
        if current_part > latest_part:
            return False

    return len(current) < len(latest)


def check_dependencies():
    missing = []
    if not EASYOCR_AVAILABLE:
        missing.append("easyocr (pip install easyocr)")
    if not PADDLEOCR_AVAILABLE:
        missing.append("paddleocr (pip install paddleocr paddlepaddle)")

    if missing:
        messagebox.showwarning(
            "Предупреждение",
            f"Следующие OCR-движки недоступны:\n{', '.join(missing)}\n\n"
            "Вы можете установить их командой:\npip install easyocr paddleocr paddlepaddle",
        )


def main():
    """Основная функция для запуска приложения."""
    try:
        read_env()
        # check_dependencies

        # Проверка замены main_new → main ДО запуска интерфейса
        # AutoUpdater.check_post_restart()
        create_interface()  # Создание интерфейса
        set_tesseract_path()

        # Обновление статуса перед проверкой обновлений
        logger.info("Проверка обновлений...")
        check_for_updates()

        updater = AutoUpdater(root)
        # updater.check_for_update()  # Здесь будет проверка наличия обновлений

        root.mainloop()  # Запуск цикла обработки событий
    except KeyboardInterrupt:
        logger.info("Программа завершена пользователем.")
        root.quit()  # Завершаем работу Tkinter, если прервано вручную


# Запуск программы
if __name__ == "__main__":
    # Настройка логгера (один раз при запуске приложения)
    logger.setup(
        log_file="main_app.log",
        gui_widget=None,  # Будет установлен позже в GUI
        max_log_size=10 * 1024 * 1024,  # 10 MB
        backup_count=5,
        log_level="DEBUG",  # 'INFO' в продакшене
    )
    logger.info("test ok")
    main()
