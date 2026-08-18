from __future__ import annotations

from types import SimpleNamespace

import openpyxl
import pytest

from pdf_ocr_app.services.excel_service import ExcelService
from pdf_ocr_app.utils.updater import AutoUpdater


def test_csv_registry_finds_start_row_by_marker(tmp_path) -> None:
    registry = tmp_path / "registry.csv"
    registry.write_text(
        "header\nheader\nheader\n1,,INV-1,prefix/MSCU1234567\n2,,INV-2,TGHU7654321\n",
        encoding="utf-8",
    )
    service = ExcelService(SimpleNamespace())

    assert service.read_registry(str(registry)) == [
        ("INV-1", "MSCU1234567"),
        ("INV-2", "TGHU7654321"),
    ]


def test_csv_registry_falls_back_to_row_5_when_no_marker(tmp_path) -> None:
    registry = tmp_path / "registry.csv"
    registry.write_text(
        "header\nheader\nheader\nheader\n,,INV-1,prefix/MSCU1234567\n,,INV-2,TGHU7654321\n",
        encoding="utf-8",
    )
    service = ExcelService(SimpleNamespace())

    assert service.read_registry(str(registry)) == [
        ("INV-1", "MSCU1234567"),
        ("INV-2", "TGHU7654321"),
    ]


def test_registry_rejects_unsupported_file_format(tmp_path) -> None:
    registry = tmp_path / "registry.txt"
    registry.write_text("data", encoding="utf-8")

    with pytest.raises(ValueError):
        ExcelService(SimpleNamespace()).read_registry(str(registry))


def test_excel_registry_finds_start_row_by_marker(tmp_path) -> None:
    registry = tmp_path / "registry.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Заголовок 1"])
    ws.append(["Заголовок 2"])
    ws.append(["Заголовок 3"])
    ws.append(["Заголовок 4"])
    ws.append([1, "", "INV-1", "prefix/MSCU1234567"])
    ws.append([2, "", "INV-2", "TGHU7654321"])
    wb.save(registry)

    service = ExcelService(SimpleNamespace())

    assert service.read_registry(str(registry)) == [
        ("INV-1", "MSCU1234567"),
        ("INV-2", "TGHU7654321"),
    ]


def test_excel_registry_finds_start_row_when_header_changes(tmp_path) -> None:
    registry = tmp_path / "registry.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Компания: ООО Ромашка"])
    ws.append(["Тип документа"])
    ws.append(["Отчёт за месяц"])
    ws.append(["№", "Пусто", "Накладная", "Контейнер"])
    ws.append([1, "", "INV-1", "prefix/MSCU1234567"])
    ws.append([2, "", "INV-2", "TGHU7654321"])
    wb.save(registry)

    service = ExcelService(SimpleNamespace())

    assert service.read_registry(str(registry)) == [
        ("INV-1", "MSCU1234567"),
        ("INV-2", "TGHU7654321"),
    ]


def test_excel_registry_skips_trailing_rows(tmp_path) -> None:
    """Хвостовые строки без данных не должны попадать в результат."""
    registry = tmp_path / "registry.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Заголовок 1"])
    ws.append(["Заголовок 2"])
    ws.append([1, "", "INV-1", "prefix/MSCU1234567"])
    ws.append([2, "", "INV-2", "TGHU7654321"])
    ws.append([3, "", "INV-3", "TRLU3456789"])
    ws.append([4, "", "INV-4", "TTNU9876543"])
    ws.append(["Итого"])  # хвостовая строка без накладной и контейнера
    ws.append([])  # пустая строка
    wb.save(registry)

    service = ExcelService(SimpleNamespace())

    assert service.read_registry(str(registry)) == [
        ("INV-1", "MSCU1234567"),
        ("INV-2", "TGHU7654321"),
        ("INV-3", "TRLU3456789"),
        ("INV-4", "TTNU9876543"),
    ]


@pytest.mark.parametrize(
    ("left", "right", "expected"),
    [
        ("2.0.5", "2.0.6", -1),
        ("2.0.6", "2.0.6", 0),
        ("2.1.0", "2.0.9", 1),
    ],
)
def test_compare_versions(left: str, right: str, expected: int) -> None:
    assert AutoUpdater.compare_versions(left, right) == expected
